# -*- coding: utf-8 -*-
"""
文件处理逻辑模块
- 负责解压/解包、Excel/CSV 文本流转换、分块复制等
- 提供将流式数据喂入 rg 的辅助方法
"""

import os
import io
import time
import shutil
import tempfile
import threading
import subprocess

from config import STREAM_CHUNK_SIZE, EXCEL_EXTS, CSV_EXTS
from utils import (
    has_cmd, popen_creationflags, emit_message_utf, emit_progress_ex,
    is_excel_file, is_csv_file
)
import process_manager as pm


def try_py7zr_extract(archive_path, dest):
    """如果可行，尝试使用 py7zr 解压 7z（以及许多其他格式）。"""
    try:
        import py7zr
    except Exception:
        return False, "py7zr not available"
    try:
        with py7zr.SevenZipFile(archive_path, mode='r') as z:
            z.extractall(path=dest)
        return True, ""
    except Exception as e:
        return False, str(e)


def try_rarfile_extract(archive_path, dest):
    """如果可用，尝试使用 Python 的 rarfile 库来解压 rar 文件。"""
    try:
        import rarfile
    except Exception:
        return False, "rarfile not available"
    try:
        rf = rarfile.RarFile(archive_path)
        rf.extractall(dest)
        return True, ""
    except Exception as e:
        return False, str(e)


def start_rg_and_feed_python_stream(rg_cmd, feed_fn):
    """
    启动 rg 进程并用 feed_fn 向其 stdin 写入解压或转换后的内容。
    feed_fn 接受一个可写二进制文件对象。
    """
    rg_proc = subprocess.Popen(
        rg_cmd,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        shell=False,
        preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None,
        creationflags=popen_creationflags()
    )

    def writer_thread():
        try:
            with rg_proc.stdin:
                if not pm.cancel_requested:
                    feed_fn(rg_proc.stdin)
        except Exception:
            try:
                rg_proc.stdin.close()
            except Exception:
                pass

    t = threading.Thread(target=writer_thread, daemon=True)
    t.start()
    return rg_proc


def python_decompress_feed(path, ext, out_stream):
    """解压单一压缩文件并写入到 out_stream（支持 gzip/bz2/lzma/lz4）。"""
    ext = (ext or '').lower()
    try:
        chunk_size = STREAM_CHUNK_SIZE
        if ext.endswith('.gz'):
            import gzip
            with gzip.open(path, 'rb') as f:
                while True:
                    if pm.cancel_requested:
                        break
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    out_stream.write(chunk)
        elif ext.endswith('.bz2'):
            import bz2
            with bz2.open(path, 'rb') as f:
                while True:
                    if pm.cancel_requested:
                        break
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    out_stream.write(chunk)
        elif ext.endswith('.xz') or ext.endswith('.txz') or ext.endswith('.lzma'):
            import lzma
            with lzma.open(path, 'rb') as f:
                while True:
                    if pm.cancel_requested:
                        break
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    out_stream.write(chunk)
        elif ext.endswith('.lz4'):
            try:
                import lz4.frame as lz4frame
                with open(path, 'rb') as raw:
                    decompressor = lz4frame.LZ4FrameDecompressor()
                    while True:
                        if pm.cancel_requested:
                            break
                        chunk = raw.read(chunk_size)
                        if not chunk:
                            break
                        out = decompressor.decompress(chunk)
                        if out:
                            out_stream.write(out)
            except Exception:
                raise
        else:
            raise RuntimeError("Unsupported python decompressor for ext: " + ext)
    except Exception:
        raise


def build_decompress_command(path_lower, real_path):
    """根据后缀构建外部解压命令（若系统支持对应工具）。"""
    path_lower = (path_lower or '').lower()
    if path_lower.endswith('.gz'):
        if has_cmd('gzip'):
            return ['gzip', '-dc', real_path]
        else:
            return None
    if path_lower.endswith('.bz2'):
        if has_cmd('bzip2'):
            return ['bzip2', '-dc', real_path]
        else:
            return None
    if path_lower.endswith('.xz') or path_lower.endswith('.txz'):
        if has_cmd('xz'):
            return ['xz', '-dc', real_path]
        else:
            return None
    if path_lower.endswith('.lz4'):
        if has_cmd('lz4'):
            return ['lz4', '-dc', real_path]
        else:
            return None
    if path_lower.endswith('.lzma'):
        if has_cmd('lzma'):
            return ['lzma', '-dc', real_path]
        else:
            return None
    if path_lower.endswith(('.7z', '.rar')):
        if has_cmd('7z'):
            return ['7z', 'x', '-so', real_path]
    return None


def list_7z_members(archive_path):
    """返回 7z 列表的成员信息（name,size），size 可能为 None。"""
    items = []
    if not has_cmd('7z'):
        return items
    try:
        p = subprocess.run(['7z', 'l', '-slt', archive_path], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, check=False)
        out_lines = p.stdout.decode('utf-8', errors='replace').splitlines()
        current_path = None
        current_type = None
        current_size = None
        for line in out_lines:
            s = line.strip()
            if not s:
                if current_path and (not current_type or current_type.lower() == 'file'):
                    items.append({'name': current_path, 'size': (int(current_size) if current_size and current_size.isdigit() else None)})
                current_path = None
                current_type = None
                current_size = None
                continue
            if s.startswith('Path = '):
                current_path = s[7:]
            elif s.startswith('Type = '):
                current_type = s[7:]
            elif s.startswith('Size = '):
                current_size = s[7:]
        if current_path and (not current_type or current_type.lower() == 'file'):
            items.append({'name': current_path, 'size': (int(current_size) if current_size and current_size.isdigit() else None)})
    except Exception:
        pass
    return items


def safe_extract_tar(tar, path):
    """安全地解包 tar，防止路径穿越。"""
    import tarfile
    for member in tar.getmembers():
        member_path = os.path.join(path, member.name)
        abs_dest = os.path.abspath(path)
        abs_target = os.path.abspath(member_path)
        if not abs_target.startswith(abs_dest + os.sep) and abs_target != abs_dest:
            raise Exception("Attempted Path Traversal in Tar File")
    tar.extractall(path)


def safe_extract_zip(zipf, path):
    """安全地解包 zip，防止路径穿越。"""
    import zipfile
    for member in zipf.namelist():
        member_path = os.path.join(path, member)
        abs_dest = os.path.abspath(path)
        abs_target = os.path.abspath(member_path)
        if not abs_target.startswith(abs_dest + os.sep) and abs_target != abs_dest:
            raise Exception("Attempted Path Traversal in Zip File")
    zipf.extractall(path)


def stream_excel_to_writer(path, out_stream):
    """
    将 xls/xlsx 文件内容以文本形式写到 out_stream（二进制写）。
    输出格式（便于 ripgrep 搜索）：
      # sheet: <sheetname>
      <cell1>\t<cell2>\t... \n
    优先使用 Python 库 openpyxl（xlsx）和 xlrd（xls）。
    """
    path_lower = (path or '').lower()
    try:
        if path_lower.endswith('.xlsx'):
            try:
                import openpyxl
            except Exception:
                emit_message_utf(f'openpyxl not installed, cannot parse xlsx: {os.path.basename(path)}\n')
                return
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sheet in wb:
                if pm.cancel_requested:
                    break
                try:
                    out_stream.write((f"# sheet: {sheet.title}\n").encode('utf-8'))
                except Exception:
                    pass
                for row in sheet.iter_rows(values_only=True):
                    try:
                        vals = []
                        for v in row:
                            if v is None:
                                vals.append('')
                            else:
                                vals.append(str(v))
                        line = '\t'.join(vals) + '\n'
                        out_stream.write(line.encode('utf-8'))
                    except Exception:
                        try:
                            safe_vals = [str(v) if v is not None else '' for v in row]
                            line = '\t'.join(safe_vals) + '\n'
                            out_stream.write(line.encode('utf-8', errors='replace'))
                        except Exception:
                            continue
            try:
                wb.close()
            except Exception:
                pass
        elif path_lower.endswith('.xls'):
            try:
                import xlrd
            except Exception:
                emit_message_utf(f'xlrd not installed, cannot parse xls: {os.path.basename(path)}\n')
                return
            wb = xlrd.open_workbook(path, on_demand=True)
            for si in range(wb.nsheets):
                sheet = wb.sheet_by_index(si)
                if pm.cancel_requested:
                    break
                try:
                    out_stream.write((f"# sheet: {sheet.name}\n").encode('utf-8'))
                except Exception:
                    pass
                for r in range(sheet.nrows):
                    try:
                        row = sheet.row_values(r)
                        vals = [(str(c) if c is not None else '') for c in row]
                        line = '\t'.join(vals) + '\n'
                        out_stream.write(line.encode('utf-8'))
                    except Exception:
                        try:
                            out_stream.write(('\t'.join([str(c) for c in sheet.row_values(r)]) + '\n').encode('utf-8', errors='replace'))
                        except Exception:
                            continue
            try:
                wb.release_resources()
            except Exception:
                pass
        else:
            emit_message_utf(f'Unsupported excel format: {path}\n')
            return
    except Exception as e:
        emit_message_utf(f'Excel parse failed for {os.path.basename(path)}: {e}\n')
        return


def stream_excel_bytes_to_writer(name_lower, data_bytes, out_stream):
    """允许传入内存字节的 Excel 转换（供归档成员或解压输出使用）。"""
    try:
        if name_lower.endswith('.xlsx'):
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
            for sheet in wb:
                if pm.cancel_requested:
                    break
                try:
                    out_stream.write((f"# sheet: {sheet.title}\n").encode('utf-8'))
                except Exception:
                    pass
                for row in sheet.iter_rows(values_only=True):
                    if pm.cancel_requested:
                        break
                    try:
                        vals = [(str(c) if c is not None else '') for c in row]
                        out_stream.write(('\t'.join(vals) + '\n').encode('utf-8'))
                    except Exception:
                        try:
                            safe_vals = [str(v) if v is not None else '' for v in row]
                            out_stream.write(('\t'.join(safe_vals) + '\n').encode('utf-8', errors='replace'))
                        except Exception:
                            continue
            try:
                wb.close()
            except Exception:
                pass
        elif name_lower.endswith('.xls'):
            import xlrd
            wb = xlrd.open_workbook(file_contents=data_bytes, on_demand=True)
            for si in range(wb.nsheets):
                sheet = wb.sheet_by_index(si)
                try:
                    out_stream.write((f"# sheet: {sheet.name}\n").encode('utf-8'))
                except Exception:
                    pass
                for r in range(sheet.nrows):
                    if pm.cancel_requested:
                        break
                    try:
                        row = sheet.row_values(r)
                        vals = [(str(c) if c is not None else '') for c in row]
                        out_stream.write(('\t'.join(vals) + '\n').encode('utf-8'))
                    except Exception:
                        try:
                            out_stream.write(('\t'.join([str(c) for c in sheet.row_values(r)]) + '\n').encode('utf-8', errors='replace'))
                        except Exception:
                            continue
            try:
                wb.release_resources()
            except Exception:
                pass
    except Exception as e:
        emit_message_utf(f'Excel parse failed (bytes) for {name_lower}: {e}\n')
        return


def stream_csv_fileobj_to_writer(fileobj, out_stream, progress_cb=None, bytes_total=None):
    """CSV 文件对象直接复制到输出（提供统一接口），自动补尾换行。"""
    try:
        last_byte = None
        done = 0
        start_ns = time.perf_counter_ns()
        while True:
            if pm.cancel_requested:
                break
            chunk = fileobj.read(64 * 1024)
            if not chunk:
                break
            out_stream.write(chunk)
            done += len(chunk)
            if progress_cb:
                elapsed_ms = int((time.perf_counter_ns() - start_ns) / 1_000_000)
                try:
                    progress_cb(done, bytes_total, elapsed_ms)
                except Exception:
                    pass
            try:
                if chunk:
                    last_byte = chunk[-1]
            except Exception:
                pass
        if last_byte is not None and last_byte != 0x0A:
            try:
                out_stream.write(b'\n')
            except Exception:
                pass
    except Exception:
        try:
            data = fileobj.read()
            if not pm.cancel_requested and data:
                out_stream.write(data)
                done += len(data)
                if progress_cb:
                    elapsed_ms = int((time.perf_counter_ns() - start_ns) / 1_000_000)
                    try:
                        progress_cb(done, bytes_total, elapsed_ms)
                    except Exception:
                        pass
                try:
                    if not pm.cancel_requested and data[-1] != 0x0A:
                        out_stream.write(b'\n')
                except Exception:
                    pass
        except Exception:
            pass


def copy_fileobj_chunked(src, dst, chunk_size: int = STREAM_CHUNK_SIZE, progress_cb=None, bytes_total=None):
    """
    分块复制 src->dst，支持可选的字节级进度回调（毫秒级计时）。
    progress_cb: callable(done_bytes:int, total_bytes:Optional[int], elapsed_ms:int)
    """
    start_ns = time.perf_counter_ns()
    done = 0
    try:
        while True:
            if pm.cancel_requested:
                try:
                    if hasattr(dst, 'flush'):
                        dst.flush()
                except Exception:
                    pass
                try:
                    if hasattr(dst, 'close'):
                        dst.close()
                except Exception:
                    pass
                break
            chunk = src.read(chunk_size)
            if not chunk:
                break
            try:
                dst.write(chunk)
                done += len(chunk)
                if progress_cb:
                    elapsed_ms = int((time.perf_counter_ns() - start_ns) / 1_000_000)
                    try:
                        progress_cb(done, bytes_total, elapsed_ms)
                    except Exception:
                        pass
            except BrokenPipeError:
                break
    except Exception:
        try:
            data = src.read()
            if not pm.cancel_requested and data:
                try:
                    dst.write(data)
                    done += len(data)
                    if progress_cb:
                        elapsed_ms = int((time.perf_counter_ns() - start_ns) / 1_000_000)
                        try:
                            progress_cb(done, bytes_total, elapsed_ms)
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception:
            pass
    try:
        if hasattr(dst, 'flush'):
            dst.flush()
    except Exception:
        pass


def spool_stream_to_temp_then_stream_excel(name_lower: str, in_stream, out_stream):
    """将输入流落盘为临时Excel文件后再流式转换（避免内存峰值）。"""
    import tempfile as _tempfile_mod
    ext = '.xlsx' if name_lower.endswith('.xlsx') else ('.xls' if name_lower.endswith('.xls') else '.xlsx')
    tmp = _tempfile_mod.NamedTemporaryFile(prefix='rg_excel_', suffix=ext, delete=False)
    tmp_path = tmp.name
    try:
        copy_fileobj_chunked(in_stream, tmp)
        try:
            tmp.flush()
        except Exception:
            pass
        try:
            tmp.close()
        except Exception:
            pass
        # 直接复用现有的按路径Excel流式转换
        stream_excel_to_writer(tmp_path, out_stream)
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
