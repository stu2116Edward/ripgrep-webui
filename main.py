from flask import Flask, render_template, request, send_from_directory, abort, jsonify
from flask_socketio import SocketIO, emit
import subprocess
import threading
import os
import signal
import time
import tempfile
import shutil
import queue
import io

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*", path='/io')

# 全局当前正在运行的 ripgrep 进程（read loop 读取这些进程的 stdout）
proc = None
# 可能存在的额外进程（例如解压器/解包器/rg），搜索/取消时需要一起清理
extra_procs = []
# 临时目录列表（例如对 archive 的解包目录），结束时需要清理
temp_dirs = []

# 取消标志（替代原先缓冲与一次性落盘逻辑）
cancel_requested = False

# 后台导出文件的流式写入句柄（按安全化后的关键字区分）
export_streams = {}

# 扩展的进度事件发射器（支持毫秒计时与字节进度）
def emit_progress_ex(phase=None, file_type=None, elapsed_ms=None,
                     bytes_done=None, bytes_total=None,
                     matches=None, files_total=None, files_done=None,
                     label=None):
    try:
        payload = {}
        if matches is not None:
            payload['matches'] = int(matches)
        if files_total is not None:
            payload['files_total'] = int(files_total)
        if files_done is not None:
            payload['files_done'] = int(files_done)
        if phase:
            payload['phase'] = str(phase)
        if file_type:
            payload['file_type'] = str(file_type)
        # elapsed_ms 已禁用：停止向前端发送时间
        if bytes_done is not None:
            try:
                bd = int(bytes_done)
                bt = int(bytes_total) if bytes_total is not None else None
                if bt is not None and bd > bt:
                    bd = bt
                payload['bytes_done'] = bd
                if bt is not None:
                    payload['bytes_total'] = bt
            except Exception:
                pass
        if label:
            payload['label'] = label
        socketio.emit('progress', payload)
    except Exception:
        # 保持后端健壮性，忽略单次发送失败
        pass

# 新增：统一 UTF-8 文本发送包装，避免乱码
def emit_message_utf(text):
    try:
        if not isinstance(text, str):
            try:
                text = str(text)
            except Exception:
                text = ''
        sanitized = text.encode('utf-8', errors='replace').decode('utf-8', errors='replace')
        socketio.emit('message', {'message': sanitized})
    except Exception:
        pass

# 关键字安全化（用于文件名前缀）
def _sanitize_keyword(keyword: str) -> str:
    try:
        safe = ''.join(c for c in (keyword or '') if c.isalnum() or c in (' ', '_', '-')).strip()
        return safe or 'search'
    except Exception:
        return 'search'

# 统一导出目录选择：固定为 /app/exports（Docker 容器卷），不再使用本地路径或环境变量
def get_exports_dir():
    try:
        os.makedirs('/app/exports', exist_ok=True)
    except Exception:
        # 在 Docker 中通常已通过挂载创建；忽略创建失败
        pass
    return '/app/exports'

# 启动一个新的导出文件写入流（每次搜索唯一文件）
def start_export_stream(safe_kw: str):
    import datetime
    try:
        # 统一导出目录选择（容器环境）
        exports_dir = get_exports_dir()
        os.makedirs(exports_dir, exist_ok=True)
        today = datetime.datetime.now().strftime('%Y-%m-%d')
        ts = int(time.time())
        filename = f"{safe_kw}_{today}_{ts}.txt"
        filepath = os.path.join(exports_dir, filename)
        fh = open(filepath, 'w', encoding='utf-8')
        export_streams[safe_kw] = {'fh': fh, 'path': filepath}
        return filepath
    except Exception:
        return None

# 追加写入文本到导出文件（若未初始化则尝试创建）
def append_export_text(safe_kw: str, text: str):
    try:
        info = export_streams.get(safe_kw)
        if not info or not info.get('fh'):
            start_export_stream(safe_kw)
            info = export_streams.get(safe_kw)
        fh = info and info.get('fh')
        if not fh:
            return
        # 确保是字符串并以 UTF-8 写入
        if not isinstance(text, str):
            try:
                text = str(text)
            except Exception:
                text = ''
        sanitized = text.encode('utf-8', errors='replace').decode('utf-8', errors='replace')
        fh.write(sanitized)
        # 实时落盘：立即刷新并尝试 fsync，确保写入不因缓冲而延迟
        try:
            fh.flush()
        except Exception:
            pass
        try:
            os.fsync(fh.fileno())
        except Exception:
            pass
    except Exception:
        pass

# 关闭指定关键字的导出流
def close_export_stream(safe_kw: str):
    try:
        info = export_streams.pop(safe_kw, None)
        if info and info.get('fh'):
            try:
                info['fh'].flush()
            except Exception:
                pass
            try:
                info['fh'].close()
            except Exception:
                pass
    except Exception:
        pass

# 关闭所有导出流（用于取消或搜索结束清理）
def close_all_export_streams():
    try:
        for safe_kw in list(export_streams.keys()):
            close_export_stream(safe_kw)
        export_streams.clear()
    except Exception:
        pass

# 支持的压缩/归档后缀（用于目录扫描时识别并单独处理）
SINGLE_COMPRESSED_EXTS = ('.gz', '.bz2', '.xz', '.lz4', '.lzma')
ARCHIVE_EXTS = (
    '.zip', '.jar', '.war',
    '.tar', '.tar.gz', '.tgz', '.tar.bz2', '.tbz2', '.tar.xz', '.txz',
    '.7z', '.rar'   # 支持 .7z / .rar
)

# 支持的电子表格/文本后缀
EXCEL_EXTS = ('.xls', '.xlsx')
CSV_EXTS = ('.csv',)

# 缓存 ripgrep 是否支持 --label 参数（None 表示尚未检测）
_RG_SUPPORTS_LABEL = None
# 进程 pid -> 自定义 label 映射（当 rg 不支持 --label 时使用）
_proc_label_map = {}


def is_single_file_compressed(filename_lower):
    if filename_lower.endswith(('.tar.gz', '.tgz', '.tar.bz2', '.tbz2', '.tar.xz', '.txz')):
        return False
    return filename_lower.endswith(SINGLE_COMPRESSED_EXTS)


def is_archive_multi_file(filename_lower):
    return filename_lower.endswith(ARCHIVE_EXTS)


def is_excel_file(filename_lower):
    return filename_lower.endswith(EXCEL_EXTS)


def is_csv_file(filename_lower):
    return filename_lower.endswith(CSV_EXTS)

# 额外的文本/数据类型扩展名
TEXT_EXTS = (
    '.txt', '.log', '.json', '.xml', '.md', '.ini', '.yaml', '.yml'
)

def classify_file_type(filename_lower: str):
    """根据文件名后缀分类文件类型：archive/compressed/excel/csv/text/other"""
    fl = filename_lower.lower()
    if is_archive_multi_file(fl):
        return 'archive'
    if is_single_file_compressed(fl):
        return 'compressed'
    if is_excel_file(fl):
        return 'excel'
    if is_csv_file(fl):
        return 'csv'
    if fl.endswith(TEXT_EXTS):
        return 'text'
    return 'other'


def strip_single_compress_ext(filename_lower):
    """去掉单一压缩扩展，返回内部真实扩展（小写，如 .xlsx/.csv），不匹配则返回空串"""
    for ce in SINGLE_COMPRESSED_EXTS:
        if filename_lower.endswith(ce):
            base = filename_lower[:-len(ce)]
            inner_ext = os.path.splitext(base)[1].lower()
            return inner_ext
    return ''


def has_cmd(name):
    """检查外部命令是否在 PATH 中"""
    return shutil.which(name) is not None


def popen_creationflags():
    """Windows: 使用 CREATE_NEW_PROCESS_GROUP 以改进终止可靠性；其他平台返回 0。"""
    try:
        if os.name == 'nt':
            import subprocess as sp
            return getattr(sp, 'CREATE_NEW_PROCESS_GROUP', 0)
        return 0
    except Exception:
        return 0


def check_rg_supports_label():
    """检测系统中 rg 的帮助输出里是否包含 --label，缓存结果"""
    global _RG_SUPPORTS_LABEL
    if _RG_SUPPORTS_LABEL is not None:
        return _RG_SUPPORTS_LABEL
    try:
        p = subprocess.run(['rg', '--help'], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, check=False)
        out = p.stdout.decode('utf-8', errors='replace')
        _RG_SUPPORTS_LABEL = ('--label' in out)
    except Exception:
        _RG_SUPPORTS_LABEL = False
    return _RG_SUPPORTS_LABEL


def try_py7zr_extract(archive_path, dest):
    """如果可行，尝试使用 py7zr 解压 7z（以及许多其他格式）"""
    try:
        import py7zr
    except Exception:
        return False, "py7zr not available"
    try:
        with py7zr.SevenZipFile(archive_path, mode='r') as z:
            z.extractall(path=dest)
        return True, ""
    except Exception as e:
        return False, str(e)


def try_rarfile_extract(archive_path, dest):
    """如果可用，尝试使用 Python 的 rarfile 库来解压 rar 文件"""
    try:
        import rarfile
    except Exception:
        return False, "rarfile not available"
    try:
        rf = rarfile.RarFile(archive_path)
        rf.extractall(dest)
        return True, ""
    except Exception as e:
        return False, str(e)


def start_rg_and_feed_python_stream(rg_cmd, feed_fn):
    """
    启动 rg 进程并用 feed_fn 向其 stdin 写入解压或转换后的内容
    feed_fn 接受一个可写二进制文件对象
    """
    rg_proc = subprocess.Popen(
        rg_cmd,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        shell=False,
        preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None,
        creationflags=popen_creationflags()
    )

    def writer_thread():
        try:
            with rg_proc.stdin:
                if not cancel_requested:
                    feed_fn(rg_proc.stdin)
        except Exception:
            try:
                rg_proc.stdin.close()
            except Exception:
                pass

    t = threading.Thread(target=writer_thread, daemon=True)
    t.start()
    return rg_proc


def python_decompress_feed(path, ext, out_stream):
    """解压单一压缩文件并写入到 out_stream（支持gzip,bz2,lzma,lz4），增加取消检查。"""
    ext = ext.lower()
    try:
        chunk_size = STREAM_CHUNK_SIZE
        if ext.endswith('.gz'):
            import gzip
            with gzip.open(path, 'rb') as f:
                while True:
                    if cancel_requested:
                        break
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    out_stream.write(chunk)
        elif ext.endswith('.bz2'):
            import bz2
            with bz2.open(path, 'rb') as f:
                while True:
                    if cancel_requested:
                        break
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    out_stream.write(chunk)
        elif ext.endswith('.xz') or ext.endswith('.txz') or ext.endswith('.lzma'):
            import lzma
            with lzma.open(path, 'rb') as f:
                while True:
                    if cancel_requested:
                        break
                    chunk = f.read(chunk_size)
                    if not chunk:
                        break
                    out_stream.write(chunk)
        elif ext.endswith('.lz4'):
            try:
                import lz4.frame as lz4frame
                with open(path, 'rb') as raw:
                    decompressor = lz4frame.LZ4FrameDecompressor()
                    while True:
                        if cancel_requested:
                            break
                        chunk = raw.read(chunk_size)
                        if not chunk:
                            break
                        out = decompressor.decompress(chunk)
                        if out:
                            out_stream.write(out)
            except Exception:
                raise
        else:
            raise RuntimeError("Unsupported python decompressor for ext: " + ext)
    except Exception:
        raise


def build_decompress_command(path_lower, real_path):
    path_lower = path_lower.lower()
    if path_lower.endswith('.gz'):
        if has_cmd('gzip'):
            return ['gzip', '-dc', real_path]
        else:
            return None
    if path_lower.endswith('.bz2'):
        if has_cmd('bzip2'):
            return ['bzip2', '-dc', real_path]
        else:
            return None
    if path_lower.endswith('.xz') or path_lower.endswith('.txz'):
        if has_cmd('xz'):
            return ['xz', '-dc', real_path]
        else:
            return None
    if path_lower.endswith('.lz4'):
        if has_cmd('lz4'):
            return ['lz4', '-dc', real_path]
        else:
            return None
    if path_lower.endswith('.lzma'):
        if has_cmd('lzma'):
            return ['lzma', '-dc', real_path]
        else:
            return None
    if path_lower.endswith(('.7z', '.rar')):
        if has_cmd('7z'):
            return ['7z', 'x', '-so', real_path]
    return None


def list_7z_members(archive_path):
    """返回 7z 列表的成员信息（name,size），size 可能为 None。"""
    items = []
    if not has_cmd('7z'):
        return items
    try:
        p = subprocess.run(['7z', 'l', '-slt', archive_path], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, check=False)
        out_lines = p.stdout.decode('utf-8', errors='replace').splitlines()
        current_path = None
        current_type = None
        current_size = None
        for line in out_lines:
            s = line.strip()
            if not s:
                if current_path and (not current_type or current_type.lower() == 'file'):
                    items.append({'name': current_path, 'size': (int(current_size) if current_size and current_size.isdigit() else None)})
                current_path = None
                current_type = None
                current_size = None
                continue
            if s.startswith('Path = '):
                current_path = s[7:]
            elif s.startswith('Type = '):
                current_type = s[7:]
            elif s.startswith('Size = '):
                current_size = s[7:]
        if current_path and (not current_type or current_type.lower() == 'file'):
            items.append({'name': current_path, 'size': (int(current_size) if current_size and current_size.isdigit() else None)})
    except Exception:
        pass
    return items


def safe_extract_tar(tar, path):
    import tarfile
    for member in tar.getmembers():
        member_path = os.path.join(path, member.name)
        abs_dest = os.path.abspath(path)
        abs_target = os.path.abspath(member_path)
        if not abs_target.startswith(abs_dest + os.sep) and abs_target != abs_dest:
            raise Exception("Attempted Path Traversal in Tar File")
    tar.extractall(path)


def safe_extract_zip(zipf, path):
    import zipfile
    for member in zipf.namelist():
        member_path = os.path.join(path, member)
        abs_dest = os.path.abspath(path)
        abs_target = os.path.abspath(member_path)
        if not abs_target.startswith(abs_dest + os.sep) and abs_target != abs_dest:
            raise Exception("Attempted Path Traversal in Zip File")
    zipf.extractall(path)


# ===== 将 Excel 文件转换为纯文本流，写入 out_stream（binary writer） =====
def stream_excel_to_writer(path, out_stream):
    """
    将 xls/xlsx 文件内容以文本形式写到 out_stream（二进制写）。
    输出格式（便于 ripgrep 搜索）：
      # sheet: <sheetname>
      <cell1>\t<cell2>\t... \n
    优先使用 Python 库 openpyxl（xlsx）和 xlrd（xls）。若缺失，会尝试通知前端并跳过该文件。
    """
    path_lower = path.lower()
    try:
        if path_lower.endswith('.xlsx'):
            try:
                import openpyxl
            except Exception:
                socketio.emit('message', {'message': f'openpyxl not installed, cannot parse xlsx: {os.path.basename(path)}\n'})
                return
            # 以只读模式打开，data_only=True 获取公式计算值（若存在）
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sheet in wb:
                if cancel_requested:
                    break
                try:
                    out_stream.write((f"# sheet: {sheet.title}\n").encode('utf-8'))
                except Exception:
                    pass
                # iter_rows(values_only=True) 返回每行的值元组
                for row in sheet.iter_rows(values_only=True):
                    try:
                        vals = []
                        for v in row:
                            if v is None:
                                vals.append('')
                            else:
                                vals.append(str(v))
                        line = '\t'.join(vals) + '\n'
                        out_stream.write(line.encode('utf-8'))
                    except Exception:
                        # 某些单元格可能编码成问题，使用 replace 处理
                        try:
                            safe_vals = [str(v) if v is not None else '' for v in row]
                            line = '\t'.join(safe_vals) + '\n'
                            out_stream.write(line.encode('utf-8', errors='replace'))
                        except Exception:
                            continue
            try:
                wb.close()
            except Exception:
                pass

        elif path_lower.endswith('.xls'):
            try:
                import xlrd
            except Exception:
                socketio.emit('message', {'message': f'xlrd not installed, cannot parse xls: {os.path.basename(path)}\n'})
                return
            wb = xlrd.open_workbook(path, on_demand=True)
            for si in range(wb.nsheets):
                sheet = wb.sheet_by_index(si)
                if cancel_requested:
                    break
                try:
                    out_stream.write((f"# sheet: {sheet.name}\n").encode('utf-8'))
                except Exception:
                    pass
                for r in range(sheet.nrows):
                    try:
                        row = sheet.row_values(r)
                        vals = [(str(c) if c is not None else '') for c in row]
                        line = '\t'.join(vals) + '\n'
                        out_stream.write(line.encode('utf-8'))
                    except Exception:
                        try:
                            out_stream.write(('\t'.join([str(c) for c in sheet.row_values(r)]) + '\n').encode('utf-8', errors='replace'))
                        except Exception:
                            continue
            try:
                wb.release_resources()
            except Exception:
                pass
        else:
            # 非支持格式
            socketio.emit('message', {'message': f'Unsupported excel format: {path}\n'})
            return
    except Exception as e:
        socketio.emit('message', {'message': f'Excel parse failed for {os.path.basename(path)}: {e}\n'})
        return

# 允许传入内存字节的 Excel 转换（供归档成员或解压输出使用）
def stream_excel_bytes_to_writer(name_lower, data_bytes, out_stream):
    try:
        if name_lower.endswith('.xlsx'):
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
            for sheet in wb:
                if cancel_requested:
                    break
                try:
                    out_stream.write((f"# sheet: {sheet.title}\n").encode('utf-8'))
                except Exception:
                    pass
                for row in sheet.iter_rows(values_only=True):
                    if cancel_requested:
                        break
                    try:
                        vals = [(str(c) if c is not None else '') for c in row]
                        out_stream.write(('\t'.join(vals) + '\n').encode('utf-8'))
                    except Exception:
                        try:
                            safe_vals = [str(v) if v is not None else '' for v in row]
                            out_stream.write(('\t'.join(safe_vals) + '\n').encode('utf-8', errors='replace'))
                        except Exception:
                            continue
            try:
                wb.close()
            except Exception:
                pass
        elif name_lower.endswith('.xls'):
            import xlrd
            wb = xlrd.open_workbook(file_contents=data_bytes, on_demand=True)
            for si in range(wb.nsheets):
                sheet = wb.sheet_by_index(si)
                try:
                    out_stream.write((f"# sheet: {sheet.name}\n").encode('utf-8'))
                except Exception:
                    pass
                for r in range(sheet.nrows):
                    if cancel_requested:
                        break
                    try:
                        row = sheet.row_values(r)
                        vals = [(str(c) if c is not None else '') for c in row]
                        out_stream.write(('\t'.join(vals) + '\n').encode('utf-8'))
                    except Exception:
                        try:
                            out_stream.write(('\t'.join([str(c) for c in sheet.row_values(r)]) + '\n').encode('utf-8', errors='replace'))
                        except Exception:
                            continue
            try:
                wb.release_resources()
            except Exception:
                pass
    except Exception as e:
        socketio.emit('message', {'message': f'Excel parse failed (bytes) for {name_lower}: {e}\n'})
        return

# CSV 文件对象直接复制到输出（提供统一接口）
def stream_csv_fileobj_to_writer(fileobj, out_stream, progress_cb=None, bytes_total=None):
    try:
        last_byte = None
        done = 0
        start_ns = time.perf_counter_ns()
        while True:
            if cancel_requested:
                break
            chunk = fileobj.read(64 * 1024)
            if not chunk:
                break
            out_stream.write(chunk)
            done += len(chunk)
            if progress_cb:
                elapsed_ms = int((time.perf_counter_ns() - start_ns) / 1_000_000)
                try:
                    progress_cb(done, bytes_total, elapsed_ms)
                except Exception:
                    pass
            try:
                if chunk:
                    last_byte = chunk[-1]
            except Exception:
                pass
        if last_byte is not None and last_byte != 0x0A:
            try:
                out_stream.write(b'\n')
            except Exception:
                pass
    except Exception:
        # 退回一次性读写，并保证行尾换行
        try:
            data = fileobj.read()
            if not cancel_requested and data:
                out_stream.write(data)
                done += len(data)
                if progress_cb:
                    elapsed_ms = int((time.perf_counter_ns() - start_ns) / 1_000_000)
                    try:
                        progress_cb(done, bytes_total, elapsed_ms)
                    except Exception:
                        pass
                try:
                    if not cancel_requested and data[-1] != 0x0A:
                        out_stream.write(b'\n')
                except Exception:
                    pass
        except Exception:
            pass

# 统一的分块拷贝（避免一次性读入内存）
STREAM_CHUNK_SIZE = 256 * 1024  # 256KB

def copy_fileobj_chunked(src, dst, chunk_size: int = STREAM_CHUNK_SIZE, progress_cb=None, bytes_total=None):
    """分块复制 src->dst，支持可选的字节级进度回调（毫秒级计时）。
    progress_cb: callable(done_bytes:int, total_bytes:Optional[int], elapsed_ms:int)
    """
    start_ns = time.perf_counter_ns()
    done = 0
    try:
        while True:
            # 取消检查：若已请求取消，主动结束写入
            if cancel_requested:
                try:
                    if hasattr(dst, 'flush'):
                        dst.flush()
                except Exception:
                    pass
                try:
                    if hasattr(dst, 'close'):
                        dst.close()
                except Exception:
                    pass
                break
            chunk = src.read(chunk_size)
            if not chunk:
                break
            try:
                dst.write(chunk)
                done += len(chunk)
                if progress_cb:
                    elapsed_ms = int((time.perf_counter_ns() - start_ns) / 1_000_000)
                    try:
                        progress_cb(done, bytes_total, elapsed_ms)
                    except Exception:
                        pass
            except BrokenPipeError:
                break
    except Exception:
        # 尝试最后一次性拷贝，尽量保证输出完整
        try:
            data = src.read()
            if not cancel_requested and data:
                try:
                    dst.write(data)
                    done += len(data)
                    if progress_cb:
                        elapsed_ms = int((time.perf_counter_ns() - start_ns) / 1_000_000)
                        try:
                            progress_cb(done, bytes_total, elapsed_ms)
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception:
            pass
    try:
        if hasattr(dst, 'flush'):
            dst.flush()
    except Exception:
        pass

# 将输入流落盘为临时Excel文件后再流式转换（避免内存峰值）
import tempfile as _tempfile_mod

def spool_stream_to_temp_then_stream_excel(name_lower: str, in_stream, out_stream):
    ext = '.xlsx' if name_lower.endswith('.xlsx') else ('.xls' if name_lower.endswith('.xls') else '.xlsx')
    tmp = _tempfile_mod.NamedTemporaryFile(prefix='rg_excel_', suffix=ext, delete=False)
    tmp_path = tmp.name
    try:
        copy_fileobj_chunked(in_stream, tmp)
        try:
            tmp.flush()
        except Exception:
            pass
        try:
            tmp.close()
        except Exception:
            pass
        # 直接复用现有的按路径Excel流式转换
        stream_excel_to_writer(tmp_path, out_stream)
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass

# ===== 结束 Excel 转换函数 =====


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/search', methods=['POST'])
def search():
    global proc, extra_procs, temp_dirs, _proc_label_map, cancel_requested
    # 只有在没有搜索正在进行时才继续
    if proc is not None:
        socketio.emit('message', {'message': 'Busy\n'})
        return "Busy"
    
    # 在开始新搜索时重置取消标志
    cancel_requested = False

    data = request.json or {}
    keyword = data.get('keyword', '')
    if not keyword:
        abort(400)
    context_before = int(data.get('context_before', 0) or 0)
    context_after = int(data.get('context_after', 0) or 0)
    file = (data.get('file') or '').strip()

    # 在提交后立即创建导出文件，确保检索过程中可实时写入，避免内存峰值
    try:
        safe_kw_init = _sanitize_keyword(keyword)
        start_export_stream(safe_kw_init)
    except Exception:
        # 保持流程继续；append_export_text 会在需要时兜底创建
        pass

    # 基本 rg 参数
    rg_base = ['rg', '-uuu', '--smart-case', '--json']

    # 确认 rg 可用，避免运行时报错
    if not has_cmd('rg'):
        socketio.emit('message', {'message': 'ripgrep 未安装或不可用，请在系统 PATH 中提供 rg。'})
        return "rg not found"

    # 上下文参数
    if context_before and context_before > 0:
        rg_base += ['-B', str(context_before)]
    if context_after and context_after > 0:
        rg_base += ['-A', str(context_after)]

    # 默认搜索路径：优先 /data，如果不存在就使用项目目录
    data_dir = '/data'
    if not os.path.isdir(data_dir):
        data_dir = os.path.dirname(__file__)

    # 根据是否指定文件构建最终检索路径
    include_glob_for_basename = None
    data_dir_abs = os.path.abspath(data_dir)
    if file:
        candidate_norm = os.path.normpath(os.path.join(data_dir_abs, file))
        candidate_abs = os.path.abspath(candidate_norm)
        # 仅允许在 data_dir 内部
        if candidate_abs.startswith(data_dir_abs + os.sep) or candidate_abs == data_dir_abs:
            # 若解析后的路径存在（文件或目录），直接使用
            if os.path.exists(candidate_abs):
                search_path = candidate_abs
            else:
                # 如果仅是文件名（没有路径分隔），在整个 data_dir 下匹配该文件名
                base_only = (os.path.basename(file) == file) and ('/' not in file) and ('\\' not in file)
                if base_only:
                    include_glob_for_basename = os.path.basename(file)
                    search_path = data_dir_abs
                else:
                    # 回退：限制在 data_dir 内并使用 basename
                    search_path = os.path.join(data_dir_abs, os.path.basename(file))
        else:
            search_path = os.path.join(data_dir_abs, os.path.basename(file))
    else:
        search_path = data_dir_abs

    # 若设置了仅文件名过滤，使用 rg 的 glob 进行包含匹配
    if include_glob_for_basename:
        rg_base += ['--glob', f'**/{include_glob_for_basename}']

    # 将要运行并监听输出的 rg 进程列表（主 rg + 为每个压缩/归档额外启动的 rg）
    all_rg_procs = []
    # 附加的系统解压/解包进程（解压器、bsdtar/unzip/7z/rg），便于 cancel 清理
    extra_procs_local = []

    # 预先计算 total_files（尽量精确）
    total_files = 0
    try:
        # 队列：用于所有 rg 进程的 stdout 行实时转发
        q = queue.Queue()

        # 前向器：读取某个 rg 进程的 stdout 并写入队列，避免 stdout 堵塞
        def forward_proc_stdout(p):
            pid = getattr(p, 'pid', id(p))
            try:
                if p.stdout is None:
                    return
                for raw in p.stdout:
                    if cancel_requested:
                        break
                    q.put((raw, pid))
            except Exception:
                pass
            finally:
                # 标识该进程的 EOF
                q.put((None, pid))
        # Helper: 启动 rg 并加入监听列表
        def start_rg_for_path(path, stdin_pipe=None, label=None, exclude_patterns=None, python_stream_feed=None):
            """
            启动 rg 以搜索 path 或从 stdin（使用 '-'）读取内容。
            当系统的 rg 不支持 --label 时，仍然接受 label 参数并把 label 存入 _proc_label_map 供后续使用。
            """
            cmd = rg_base.copy()
            if exclude_patterns:
                for pat in exclude_patterns:
                    cmd += ['--glob', f'!{pat}']
            supports_label = check_rg_supports_label()
            # 当通过 stdin 流式喂入时，确保将二进制当作文本处理（避免被跳过）
            if path == '-':
                cmd += ['-a']
            if label and supports_label:
                # rg 支持 --label：把 label 传给 rg（此时 rg 会在 JSON 的 begin/path 中显示 label）
                cmd += ['--label', label, '--', keyword, '-'] if path == '-' else ['--label', label, '--', keyword, path]
            else:
                # rg 不支持 --label：不传此参数，仍用 stdin 或 path 搜索
                if path == '-':
                    cmd += ['--', keyword, '-']
                else:
                    cmd += ['--', keyword, path]

            if python_stream_feed:
                rg_proc = start_rg_and_feed_python_stream(cmd, python_stream_feed)
            else:
                rg_proc = subprocess.Popen(
                    cmd,
                    stdin=stdin_pipe,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    shell=False,
                    preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                )

            # 如果 rg 不支持 --label 且用户希望有 label，保存映射以便后续在解析时使用
            if label and not supports_label:
                try:
                    pid = rg_proc.pid
                    if pid is not None:
                        _proc_label_map[pid] = label
                except Exception:
                    pass
            # 立即启动该 rg 的 stdout 前向线程，避免在长时间 stdin 写入时发生阻塞
            try:
                t = threading.Thread(target=forward_proc_stdout, args=(rg_proc,), daemon=True)
                t.start()
            except Exception:
                pass

            all_rg_procs.append(rg_proc)
            # 注册到清理列表，确保取消时一并终止
            try:
                extra_procs_local.append(rg_proc)
                extra_procs.append(rg_proc)
            except Exception:
                pass
            return rg_proc

        # 如果用户指定了一个已存在的单个文件 -> 与以前一样处理，但使用 Python 回退
        if file and os.path.isfile(search_path):
            file_lower = file.lower()

            # 支持 Excel 文件（xls/xlsx）直接流式检索
            if is_excel_file(file_lower):
                total_files = 1
                safe_label = os.path.basename(file)
                # 使用 python 库将 excel 转为文本流并传给 rg
                def feed_fn(w, p=search_path):
                    stream_excel_to_writer(p, w)

                # 启动 rg（从 stdin 读取）
                try:
                    start_rg_for_path('-', label=safe_label, python_stream_feed=feed_fn)
                except Exception as e:
                    socketio.emit('message', {'message': f'Failed to start rg for excel: {e}\n'})
                    return "Excel stream failed"

            # 处理单文件压缩：解压并按内部类型转换后流式喂入 rg
            elif is_single_file_compressed(file_lower):
                try:
                    rel_label = os.path.basename(file)
                    dec_cmd = build_decompress_command(file_lower, search_path)
                    inner_lower = strip_single_compress_ext(file_lower)
                    if dec_cmd:
                        decompressor_proc = subprocess.Popen(
                            dec_cmd,
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE,
                            shell=False,
                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                        )
                        extra_procs_local.append(decompressor_proc)
                        extra_procs.append(decompressor_proc)
                        if is_excel_file(inner_lower):
                            def feed_fn(w, proc=decompressor_proc, name=inner_lower, lb=rel_label):
                                def _cb(done, total, elapsed_ms):
                                    emit_progress_ex(
                                        phase='decompress', file_type='compressed',
                                        elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                        label=lb
                                    )
                                # 直接将解压流落盘到临时文件，再流式转换为文本，避免内存峰值
                                spool_stream_to_temp_then_stream_excel(name, proc.stdout, w)
                            start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                        else:
                            def feed_fn(w, proc=decompressor_proc, lb=rel_label):
                                size_c = None
                                try:
                                    size_c = os.path.getsize(search_path)
                                except Exception:
                                    size_c = None
                                def _cb(done, total, elapsed_ms):
                                    emit_progress_ex(
                                        phase='decompress', file_type='compressed',
                                        elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                        label=lb
                                    )
                                copy_fileobj_chunked(proc.stdout, w, progress_cb=_cb, bytes_total=size_c)
                            start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                    else:
                        # 使用 python 解压
                        if is_excel_file(inner_lower):
                            def feed_fn(w, p=search_path, e=file_lower, inner=inner_lower):
                                # 将解压数据直接落盘到临时文件，避免一次性驻留内存
                                tmp = _tempfile_mod.NamedTemporaryFile(prefix='rg_dec_excel_', suffix=('.xlsx' if inner.endswith('.xlsx') else '.xls'), delete=False)
                                tmp_path = tmp.name
                                try:
                                    try:
                                        python_decompress_feed(p, e, tmp)
                                    finally:
                                        try:
                                            tmp.flush()
                                        except Exception:
                                            pass
                                        try:
                                            tmp.close()
                                        except Exception:
                                            pass
                                    stream_excel_to_writer(tmp_path, w)
                                finally:
                                    try:
                                        os.remove(tmp_path)
                                    except Exception:
                                        pass
                            start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                        else:
                            def feed_fn(w, p=search_path, e=file_lower):
                                python_decompress_feed(p, e, w)
                            start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                except Exception as e:
                    socketio.emit('message', {'message': f'Single-compressed stream failed: {e}\n'})
                    return "Single-compressed stream failed"

            # 直接流式处理 tar 系列归档（gz/bz2/xz/无压缩）
            elif file_lower.endswith(('.tar.gz', '.tgz', '.tar.bz2', '.tbz2', '.tar.xz', '.txz', '.tar')):
                # 直接流式处理 tar 归档内容，无需落盘；为每个归档内文件启动一个 rg（stdin）
                try:
                    import tarfile
                    if file_lower.endswith(('.tar.gz', '.tgz')):
                        tar_mode = 'r:gz'
                    elif file_lower.endswith(('.tar.bz2', '.tbz2')):
                        tar_mode = 'r:bz2'
                    elif file_lower.endswith(('.tar.xz', '.txz')):
                        tar_mode = 'r:xz'
                    else:
                        tar_mode = 'r'
                    total_files = 0
                    with tarfile.open(search_path, mode=tar_mode) as tar:
                        for member in tar.getmembers():
                            if member.isfile():
                                total_files += 1
                                f = tar.extractfile(member)
                                if f:
                                    label = f"{os.path.basename(file)}/{member.name}"
                                    cmd = rg_base.copy()
                                    supports_label = check_rg_supports_label()
                                    if supports_label:
                                        cmd += ['--label', label, '--', keyword, '-']
                                    else:
                                        cmd += ['--', keyword, '-']
                                    # 启动 rg 并把归档内文件内容写入其 stdin
                                    try:
                                        rg_proc = subprocess.Popen(
                                            cmd,
                                            stdin=subprocess.PIPE,
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.STDOUT,
                                            shell=False,
                                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None,
                                            creationflags=popen_creationflags()
                                        )
                                        extra_procs_local.append(rg_proc)
                                        all_rg_procs.append(rg_proc)
                                        if not supports_label:
                                            try:
                                                _proc_label_map[rg_proc.pid] = label
                                            except Exception:
                                                pass
                                        # 将文件流写入 rg stdin（阻塞写入是可以的，因为这是独立进程）
                                        try:
                                            lower = member.name.lower()
                                            if is_excel_file(lower):
                                                data = f.read()
                                                stream_excel_bytes_to_writer(lower, data, rg_proc.stdin)
                                            elif is_csv_file(lower):
                                                stream_csv_fileobj_to_writer(f, rg_proc.stdin)
                                            else:
                                                shutil.copyfileobj(f, rg_proc.stdin)
                                        except Exception:
                                            pass
                                        try:
                                            rg_proc.stdin.close()
                                        except Exception:
                                            pass
                                    except Exception:
                                        # 如果启动 rg 失败，跳过该文件
                                        continue
                    # 如果归档为空，仍要保证至少一个进程以触发“无结果”逻辑
                    if total_files == 0:
                        total_files = 1
                        start_rg_for_path(search_path)
                except Exception as e:
                    socketio.emit('message', {'message': f'Tar stream failed: {e}\n'})
                    return "Tar stream failed"

            # 直接流式处理 zip/rar/7z 等归档（无外部 extractor 情况下的回退）
            elif is_archive_multi_file(file_lower):
                # 优先尝试 python 库的“逐文件流式读取并喂入 rg”方式，避免依赖外部 extractor
                streamed = False
                try:
                    # zip
                    if file_lower.endswith(('.zip', '.jar', '.war')):
                        import zipfile
                        with zipfile.ZipFile(search_path, 'r') as zf:
                            info_list = [zi for zi in zf.infolist() if not (hasattr(zi, 'is_dir') and zi.is_dir()) and not zi.filename.endswith('/')]
                            total_files = 0
                            for info in info_list:
                                name = info.filename
                                size = getattr(info, 'file_size', None)
                                total_files += 1
                                label = f"{os.path.basename(file)}/{name}"
                                supports_label = check_rg_supports_label()
                                if supports_label:
                                    cmd = rg_base.copy() + ['-a', '--label', label, '--', keyword, '-']
                                else:
                                    cmd = rg_base.copy() + ['-a', '--', keyword, '-']
                                try:
                                    rg_proc = subprocess.Popen(
                                        cmd,
                                        stdin=subprocess.PIPE,
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.STDOUT,
                                        shell=False,
                                        preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None,
                                        creationflags=popen_creationflags()
                                    )
                                    # 立即启动 stdout 前向线程，避免阻塞
                                    try:
                                        t = threading.Thread(target=forward_proc_stdout, args=(rg_proc,), daemon=True)
                                        t.start()
                                    except Exception:
                                        pass
                                    extra_procs_local.append(rg_proc)
                                    all_rg_procs.append(rg_proc)
                                    try:
                                        extra_procs.append(rg_proc)
                                    except Exception:
                                        pass
                                    if not supports_label:
                                        try:
                                            _proc_label_map[rg_proc.pid] = label
                                        except Exception:
                                            pass
                                    with zf.open(name, 'r') as member_f:
                                        try:
                                            lower = name.lower()
                                            # 解压阶段进度：基于未压缩字节大小（如可用）
                                            def _cb(done, total, elapsed_ms):
                                                emit_progress_ex(
                                                    phase='decompress', file_type='archive',
                                                    elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                                    label=label
                                                )
                                            if is_excel_file(lower):
                                                spool_stream_to_temp_then_stream_excel(lower, member_f, rg_proc.stdin)
                                            elif is_csv_file(lower):
                                                stream_csv_fileobj_to_writer(member_f, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                            else:
                                                copy_fileobj_chunked(member_f, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                        except Exception:
                                            pass
                                    try:
                                        rg_proc.stdin.close()
                                    except Exception:
                                        pass
                                    # 串行化：等待当前成员的检索完成后再继续下一个
                                    try:
                                        rg_proc.wait()
                                    except Exception:
                                        pass
                                except Exception:
                                    continue
                            streamed = True
                    # rar
                    elif file_lower.endswith('.rar'):
                        streamed_local = False
                        # 首先尝试使用 rarfile 库进行逐成员流式处理
                        try:
                            import rarfile
                            rf = rarfile.RarFile(search_path)
                            members = rf.infolist()
                            total_files = 0
                            for mi in members:
                                if mi.isdir():
                                    continue
                                total_files += 1
                                name = mi.filename
                                label = f"{os.path.basename(file)}/{name}"
                                supports_label = check_rg_supports_label()
                                if supports_label:
                                    cmd = rg_base.copy() + ['-a', '--label', label, '--', keyword, '-']
                                else:
                                    cmd = rg_base.copy() + ['-a', '--', keyword, '-']
                                try:
                                    rg_proc = subprocess.Popen(
                                        cmd,
                                        stdin=subprocess.PIPE,
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.STDOUT,
                                        shell=False,
                                        preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                    )
                                    # 立即启动 stdout 前向线程，保持实时输出
                                    try:
                                        t = threading.Thread(target=forward_proc_stdout, args=(rg_proc,), daemon=True)
                                        t.start()
                                    except Exception:
                                        pass
                                    extra_procs_local.append(rg_proc)
                                    all_rg_procs.append(rg_proc)
                                    try:
                                        extra_procs.append(rg_proc)
                                    except Exception:
                                        pass
                                    if not supports_label:
                                        try:
                                            _proc_label_map[rg_proc.pid] = label
                                        except Exception:
                                            pass
                                    with rf.open(mi) as member_f:
                                        try:
                                            lower = name.lower()
                                            size = getattr(mi, 'file_size', None)
                                            def _cb(done, total, elapsed_ms):
                                                emit_progress_ex(
                                                    phase='decompress', file_type='archive',
                                                    elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                                    label=label
                                                )
                                            if is_excel_file(lower):
                                                spool_stream_to_temp_then_stream_excel(lower, member_f, rg_proc.stdin)
                                            elif is_csv_file(lower):
                                                stream_csv_fileobj_to_writer(member_f, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                            else:
                                                copy_fileobj_chunked(member_f, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                        except Exception:
                                            pass
                                    try:
                                        rg_proc.stdin.close()
                                    except Exception:
                                        pass
                                    # 串行化等待当前成员的检索结束
                                    try:
                                        rg_proc.wait()
                                    except Exception:
                                        pass
                                except Exception:
                                    continue
                            streamed_local = True
                        except Exception:
                            streamed_local = False
                        # 回退：使用 7z 对 .rar 进行无落盘逐成员流式处理
                        if not streamed_local and has_cmd('7z'):
                            try:
                                members = list_7z_members(search_path)
                                total_files = 0
                                for m in members:
                                    name = m.get('name')
                                    size = m.get('size')
                                    total_files += 1
                                    label = f"{os.path.basename(file)}/{name}"
                                    supports_label = check_rg_supports_label()
                                    if supports_label:
                                        cmd = rg_base.copy() + ['-a', '--label', label, '--', keyword, '-']
                                    else:
                                        cmd = rg_base.copy() + ['-a', '--', keyword, '-']
                                    try:
                                        rg_proc = subprocess.Popen(
                                            cmd,
                                            stdin=subprocess.PIPE,
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.STDOUT,
                                            shell=False,
                                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                        )
                                        # 立即启动 stdout 前向线程
                                        try:
                                            t = threading.Thread(target=forward_proc_stdout, args=(rg_proc,), daemon=True)
                                            t.start()
                                        except Exception:
                                            pass
                                        extra_procs_local.append(rg_proc)
                                        all_rg_procs.append(rg_proc)
                                        try:
                                            extra_procs.append(rg_proc)
                                        except Exception:
                                            pass
                                        if not supports_label:
                                            try:
                                                _proc_label_map[rg_proc.pid] = label
                                            except Exception:
                                                pass
                                        dec_proc = subprocess.Popen(
                                            ['7z', 'x', '-so', search_path, name],
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            shell=False,
                                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None,
                                            creationflags=popen_creationflags()
                                        )
                                        extra_procs_local.append(dec_proc)
                                        try:
                                            lower = name.lower()
                                            def _cb(done, total, elapsed_ms):
                                                emit_progress_ex(
                                                    phase='decompress', file_type='archive',
                                                    elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                                    label=label
                                                )
                                            if is_excel_file(lower):
                                                spool_stream_to_temp_then_stream_excel(lower, dec_proc.stdout, rg_proc.stdin)
                                            elif is_csv_file(lower):
                                                stream_csv_fileobj_to_writer(dec_proc.stdout, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                            else:
                                                copy_fileobj_chunked(dec_proc.stdout, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                        except Exception:
                                            pass
                                        try:
                                            rg_proc.stdin.close()
                                        except Exception:
                                            pass
                                        try:
                                            dec_proc.stdout.close()
                                        except Exception:
                                            pass
                                        try:
                                            dec_proc.wait()
                                        except Exception:
                                            pass
                                        # 串行化：等待 rg 完成当前成员的检索
                                        try:
                                            rg_proc.wait()
                                        except Exception:
                                            pass
                                    except Exception:
                                        continue
                                streamed_local = True
                            except Exception:
                                streamed_local = False
                        streamed = streamed_local
                    # 7z 优先使用 7z CLI 逐成员流式；若不可用则回退 py7zr
                    elif file_lower.endswith('.7z'):
                        streamed_local = False
                        if has_cmd('7z'):
                            try:
                                members = list_7z_members(search_path)
                                total_files = 0
                                for m in members:
                                    name = m.get('name')
                                    size = m.get('size')
                                    total_files += 1
                                    label = f"{os.path.basename(file)}/{name}"
                                    supports_label = check_rg_supports_label()
                                    cmd = rg_base.copy() + (['-a', '--label', label, '--', keyword, '-'] if supports_label else ['-a', '--', keyword, '-'])
                                    try:
                                        rg_proc = subprocess.Popen(
                                            cmd,
                                            stdin=subprocess.PIPE,
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.STDOUT,
                                            shell=False,
                                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                        )
                                        # 启动 stdout 前向线程
                                        try:
                                            t = threading.Thread(target=forward_proc_stdout, args=(rg_proc,), daemon=True)
                                            t.start()
                                        except Exception:
                                            pass
                                        extra_procs_local.append(rg_proc)
                                        all_rg_procs.append(rg_proc)
                                        try:
                                            extra_procs.append(rg_proc)
                                        except Exception:
                                            pass
                                        if not supports_label:
                                            try:
                                                _proc_label_map[rg_proc.pid] = label
                                            except Exception:
                                                pass
                                        dec_proc = subprocess.Popen(
                                            ['7z', 'x', '-so', search_path, name],
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            shell=False,
                                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                        )
                                        extra_procs_local.append(dec_proc)
                                        try:
                                            lower = name.lower()
                                            def _cb(done, total, elapsed_ms):
                                                emit_progress_ex(
                                                    phase='decompress', file_type='archive',
                                                    elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                                    label=label
                                                )
                                            if is_excel_file(lower):
                                                spool_stream_to_temp_then_stream_excel(lower, dec_proc.stdout, rg_proc.stdin)
                                            elif is_csv_file(lower):
                                                stream_csv_fileobj_to_writer(dec_proc.stdout, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                            else:
                                                copy_fileobj_chunked(dec_proc.stdout, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                        except Exception:
                                            pass
                                        try:
                                            rg_proc.stdin.close()
                                        except Exception:
                                            pass
                                        try:
                                            dec_proc.stdout.close()
                                        except Exception:
                                            pass
                                        try:
                                            dec_proc.wait()
                                        except Exception:
                                            pass
                                        # 串行化：等待 rg 完成本成员检索
                                        try:
                                            rg_proc.wait()
                                        except Exception:
                                            pass
                                    except Exception:
                                        continue
                                streamed_local = True
                            except Exception:
                                streamed_local = False
                        if not streamed_local:
                            try:
                                import py7zr
                                with py7zr.SevenZipFile(search_path, mode='r') as z:
                                    try:
                                        names = [n for n in z.getnames() if n]
                                    except Exception:
                                        try:
                                            names = list((z.readall() or {}).keys())
                                        except Exception:
                                            names = []
                                    total_files = 0
                                    for name in names:
                                        total_files += 1
                                        label = f"{os.path.basename(file)}/{name}"
                                        supports_label = check_rg_supports_label()
                                        cmd = rg_base.copy() + (['-a', '--label', label, '--', keyword, '-'] if supports_label else ['-a', '--', keyword, '-'])
                                        try:
                                            rg_proc = subprocess.Popen(
                                                cmd,
                                                stdin=subprocess.PIPE,
                                                stdout=subprocess.PIPE,
                                                stderr=subprocess.STDOUT,
                                                shell=False,
                                                preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                            )
                                            # 启动 stdout 前向线程
                                            try:
                                                t = threading.Thread(target=forward_proc_stdout, args=(rg_proc,), daemon=True)
                                                t.start()
                                            except Exception:
                                                pass
                                            extra_procs_local.append(rg_proc)
                                            all_rg_procs.append(rg_proc)
                                            try:
                                                extra_procs.append(rg_proc)
                                            except Exception:
                                                pass
                                            if not supports_label:
                                                try:
                                                    _proc_label_map[rg_proc.pid] = label
                                                except Exception:
                                                    pass
                                            try:
                                                dm = z.read([name])
                                                obj = dm.get(name)
                                                if obj is None:
                                                    raise Exception('py7zr read returned None')
                                                lower = name.lower()
                                                if is_excel_file(lower):
                                                    # 先落盘再流式
                                                    bio = io.BytesIO(obj if isinstance(obj, (bytes, bytearray)) else (obj.getvalue() if hasattr(obj, 'getvalue') else obj.read()))
                                                    spool_stream_to_temp_then_stream_excel(lower, bio, rg_proc.stdin)
                                                elif is_csv_file(lower):
                                                    if isinstance(obj, (bytes, bytearray)):
                                                        data_bytes = bytes(obj)
                                                        rg_proc.stdin.write(data_bytes)
                                                        if len(data_bytes) == 0 or data_bytes[-1] != 0x0A:
                                                            rg_proc.stdin.write(b'\n')
                                                    else:
                                                        stream_csv_fileobj_to_writer(obj, rg_proc.stdin)
                                                else:
                                                    if isinstance(obj, (bytes, bytearray)):
                                                        bio = io.BytesIO(bytes(obj))
                                                        copy_fileobj_chunked(bio, rg_proc.stdin)
                                                    else:
                                                        copy_fileobj_chunked(obj, rg_proc.stdin)
                                            except Exception:
                                                pass
                                            try:
                                                rg_proc.stdin.close()
                                            except Exception:
                                                pass
                                            # 串行化：等待 rg 完成本成员检索
                                            try:
                                                rg_proc.wait()
                                            except Exception:
                                                pass
                                        except Exception:
                                            continue
                                streamed_local = True
                            except Exception:
                                streamed_local = False
                        streamed = streamed_local
                except Exception:
                    streamed = False

                if not streamed:
                    # 若 python 库流式处理失败，再退回到原来的“解包到临时目录”的策略（如有 unzip/7z/bsdtar）
                    temp_dir_for_archive = tempfile.mkdtemp(prefix='rg_archive_')
                    temp_dirs.append(temp_dir_for_archive)
                    extracted_ok = False
                    try:
                        import tarfile, zipfile
                        if file_lower.endswith(('.zip', '.jar', '.war')):
                            with zipfile.ZipFile(search_path, 'r') as zf:
                                safe_extract_zip(zf, temp_dir_for_archive)
                            extracted_ok = True
                        else:
                            # tar/other handled by tarfile
                            with tarfile.open(search_path, 'r:*') as tf:
                                safe_extract_tar(tf, temp_dir_for_archive)
                            extracted_ok = True
                    except Exception:
                        extracted_ok = False

                    if not extracted_ok:
                        if file_lower.endswith('.7z'):
                            ok, msg = try_py7zr_extract(search_path, temp_dir_for_archive)
                            if ok:
                                extracted_ok = True
                        elif file_lower.endswith('.rar'):
                            ok, msg = try_rarfile_extract(search_path, temp_dir_for_archive)
                            if ok:
                                extracted_ok = True

                    if not extracted_ok:
                        try:
                            if file_lower.endswith(('.zip', '.jar', '.war')):
                                extract_cmd = ['unzip', '-qq', search_path, '-d', temp_dir_for_archive]
                            elif file_lower.endswith(('.7z', '.rar')):
                                if has_cmd('7z'):
                                    extract_cmd = ['7z', 'x', '-y', search_path, f'-o{temp_dir_for_archive}']
                                else:
                                    raise FileNotFoundError('7z not available')
                            else:
                                extract_cmd = ['bsdtar', '-xf', search_path, '-C', temp_dir_for_archive]
                            extract_proc = subprocess.Popen(
                                extract_cmd,
                                stdout=subprocess.PIPE,
                                stderr=subprocess.PIPE,
                                shell=False,
                                preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None,
                                creationflags=popen_creationflags()
                            )
                            extra_procs_local.append(extract_proc)
                            extra_procs.append(extract_proc)
                            out, err = extract_proc.communicate()
                            if extract_proc.returncode != 0:
                                msg = err.decode('utf-8', errors='replace') if err else 'Unknown extraction error'
                                socketio.emit('message', {'message': f'Extraction failed: {msg}\n'})
                                try:
                                    shutil.rmtree(temp_dir_for_archive)
                                except Exception:
                                    pass
                                temp_dirs.remove(temp_dir_for_archive)
                                return "Extraction failed"
                        except FileNotFoundError:
                            socketio.emit('message', {'message': f'Missing extractor on system for archive: {file}\n'})
                            try:
                                shutil.rmtree(temp_dir_for_archive)
                            except Exception:
                                pass
                            temp_dirs.remove(temp_dir_for_archive)
                            return "Missing extractor"

                    # 解包后按目录启动 rg（同时为 Excel 文件单独流式转换）
                    excel_files_list = []
                    non_excel_count = 0
                    for root, _, fns in os.walk(temp_dir_for_archive):
                        for fn in fns:
                            lower = fn.lower()
                            full = os.path.join(root, fn)
                            if is_excel_file(lower):
                                excel_files_list.append(full)
                            else:
                                non_excel_count += 1
                    total_files = (non_excel_count + len(excel_files_list)) if (non_excel_count + len(excel_files_list)) > 0 else 1
                    # 目录扫描时排除 Excel，避免重复
                    exclude_patterns = [f'**/*{ext}' for ext in EXCEL_EXTS]
                    # 串行化：目录扫描结束前不并行其他任务
                    _p_dir = start_rg_for_path(temp_dir_for_archive, exclude_patterns=exclude_patterns)
                    try:
                        _p_dir.wait()
                    except Exception:
                        pass
                    # 对每个 Excel 文件做流式转换并交给 rg
                    for full in excel_files_list:
                        # label 使用 归档名/归档内相对路径
                        rel_inside = os.path.relpath(full, temp_dir_for_archive)
                        label = f"{os.path.basename(file)}/{rel_inside}"
                        def feed_fn(w, p=full):
                            stream_excel_to_writer(p, w)
                        try:
                            _p_excel = start_rg_for_path('-', label=label, python_stream_feed=feed_fn)
                            try:
                                _p_excel.wait()
                            except Exception:
                                pass
                        except Exception:
                            continue

            else:
                # 单个普通文件：计为1
                total_files = 1
                # 对于非压缩/归档/Excel 的单文件，使用 stdin 流式喂入并按字节上报进度
                try:
                    rel_label = os.path.relpath(search_path, data_dir)
                except Exception:
                    rel_label = os.path.basename(search_path)
                file_lower = file.lower()
                if (not is_single_file_compressed(file_lower)) and (not is_archive_multi_file(file_lower)) and (not is_excel_file(file_lower)):
                    def feed_fn(w, p=search_path, lb=rel_label, fl=file_lower):
                        size_c = None
                        try:
                            size_c = os.path.getsize(p)
                        except Exception:
                            size_c = None
                        ft = classify_file_type(fl)
                        def _cb(done, total, elapsed_ms, ft_local=ft, lb_local=lb):
                            emit_progress_ex(
                                phase='scan', file_type=ft_local,
                                elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                label=lb_local
                            )
                        try:
                            with open(p, 'rb') as f:
                                copy_fileobj_chunked(f, w, progress_cb=_cb, bytes_total=size_c)
                        except Exception:
                            pass
                    _p_single = start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                    try:
                        _p_single.wait()
                    except Exception:
                        pass
                else:
                    _p_path = start_rg_for_path(search_path)
                    try:
                        _p_path.wait()
                    except Exception:
                        pass
        else:
            # 目录搜索（默认）-> 包括压缩文件、归档和 Excel
            regular_files = []
            compressed_files = []
            archive_files = []
            excel_files = []
            # 先收集文件，这样我们才能准确计算
            for root, _, fns in os.walk(search_path):
                for fn in fns:
                    # 若启用了仅文件名过滤，则只统计匹配的文件
                    if include_glob_for_basename and os.path.basename(fn) != include_glob_for_basename:
                        continue
                    full = os.path.join(root, fn)
                    fn_lower = fn.lower()
                    if is_archive_multi_file(fn_lower):
                        archive_files.append(full)
                    elif is_single_file_compressed(fn_lower):
                        compressed_files.append(full)
                    elif is_excel_file(fn_lower):
                        excel_files.append(full)
                    else:
                        regular_files.append(full)
            total_files = len(regular_files) + len(compressed_files) + len(excel_files)
            # 对于存档，首先尝试使用 Python 库进行流式处理；如果流式处理成功，它将针对每个成员启动 rg
            for full_archive in archive_files:
                streamed = False
                try:
                    alc = full_archive.lower()
                    # zip streaming
                    if alc.endswith(('.zip', '.jar', '.war')):
                        try:
                            import zipfile
                            with zipfile.ZipFile(full_archive, 'r') as zf:
                                info_list = [zi for zi in zf.infolist() if not (hasattr(zi, 'is_dir') and zi.is_dir()) and not zi.filename.endswith('/')]
                                if info_list:
                                    for info in info_list:
                                        name = info.filename
                                        size = getattr(info, 'file_size', None)
                                        label = os.path.relpath(full_archive, data_dir) + '/' + name
                                        supports_label = check_rg_supports_label()
                                        if supports_label:
                                            cmd = rg_base.copy() + ['-a', '--label', label, '--', keyword, '-']
                                        else:
                                            cmd = rg_base.copy() + ['-a', '--', keyword, '-']
                                        try:
                                            rg_proc = subprocess.Popen(
                                                cmd,
                                                stdin=subprocess.PIPE,
                                                stdout=subprocess.PIPE,
                                                stderr=subprocess.STDOUT,
                                                shell=False,
                                                preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                            )
                                            extra_procs_local.append(rg_proc)
                                            all_rg_procs.append(rg_proc)
                                            if not supports_label:
                                                _proc_label_map[rg_proc.pid] = label
                                            with zf.open(name, 'r') as member_f:
                                                try:
                                                    lower = name.lower()
                                                    def _cb(done, total, elapsed_ms):
                                                        emit_progress_ex(
                                                            phase='decompress', file_type='archive',
                                                            elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                                            label=label
                                                        )
                                                    if is_excel_file(lower):
                                                        spool_stream_to_temp_then_stream_excel(lower, member_f, rg_proc.stdin)
                                                    elif is_csv_file(lower):
                                                        stream_csv_fileobj_to_writer(member_f, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                                    else:
                                                        copy_fileobj_chunked(member_f, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                                except Exception:
                                                    pass
                                            try:
                                                rg_proc.stdin.close()
                                            except Exception:
                                                pass
                                            # 串行化：等待 rg 完成本成员检索
                                            try:
                                                rg_proc.wait()
                                            except Exception:
                                                pass
                                            total_files += 1
                                        except Exception:
                                            continue
                                    streamed = True
                        except Exception:
                            streamed = False
                    # rar streaming
                    elif alc.endswith('.rar'):
                        streamed_local = False
                        # 优先尝试 rarfile 库进行逐成员流式处理
                        try:
                            import rarfile
                            rf = rarfile.RarFile(full_archive)
                            members = rf.infolist()
                            for mi in members:
                                if mi.isdir():
                                    continue
                                name = mi.filename
                                label = os.path.relpath(full_archive, data_dir) + '/' + name
                                supports_label = check_rg_supports_label()
                                if supports_label:
                                    cmd = rg_base.copy() + ['-a', '--label', label, '--', keyword, '-']
                                else:
                                    cmd = rg_base.copy() + ['-a', '--', keyword, '-']
                                try:
                                    rg_proc = subprocess.Popen(
                                        cmd,
                                        stdin=subprocess.PIPE,
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.STDOUT,
                                        shell=False,
                                        preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                    )
                                    extra_procs_local.append(rg_proc)
                                    all_rg_procs.append(rg_proc)
                                    if not supports_label:
                                        _proc_label_map[rg_proc.pid] = label
                                    with rf.open(mi) as member_f:
                                        try:
                                            lower = name.lower()
                                            size = getattr(mi, 'file_size', None)
                                            def _cb(done, total, elapsed_ms):
                                                emit_progress_ex(
                                                    phase='decompress', file_type='archive',
                                                    elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                                    label=label
                                                )
                                            if is_excel_file(lower):
                                                spool_stream_to_temp_then_stream_excel(lower, member_f, rg_proc.stdin)
                                            elif is_csv_file(lower):
                                                stream_csv_fileobj_to_writer(member_f, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                            else:
                                                copy_fileobj_chunked(member_f, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                        except Exception:
                                            pass
                                    try:
                                        rg_proc.stdin.close()
                                    except Exception:
                                        pass
                                    # 串行化：等待当前成员检索完成
                                    try:
                                        rg_proc.wait()
                                    except Exception:
                                        pass
                                    total_files += 1
                                except Exception:
                                    continue
                            streamed_local = True
                        except Exception:
                            streamed_local = False
                        # 回退：使用 7z 对 .rar 进行无落盘逐成员流式处理
                        if not streamed_local and has_cmd('7z'):
                            try:
                                members = list_7z_members(full_archive)
                                for m in members:
                                    name = m.get('name')
                                    size = m.get('size')
                                    label = os.path.relpath(full_archive, data_dir) + '/' + name
                                    supports_label = check_rg_supports_label()
                                    if supports_label:
                                        cmd = rg_base.copy() + ['-a', '--label', label, '--', keyword, '-']
                                    else:
                                        cmd = rg_base.copy() + ['-a', '--', keyword, '-']
                                    try:
                                        rg_proc = subprocess.Popen(
                                            cmd,
                                            stdin=subprocess.PIPE,
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.STDOUT,
                                            shell=False,
                                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                        )
                                        extra_procs_local.append(rg_proc)
                                        all_rg_procs.append(rg_proc)
                                        if not supports_label:
                                            _proc_label_map[rg_proc.pid] = label
                                        dec_proc = subprocess.Popen(
                                            ['7z', 'x', '-so', full_archive, name],
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            shell=False,
                                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None,
                                            creationflags=popen_creationflags()
                                        )
                                        extra_procs_local.append(dec_proc)
                                        try:
                                            lower = name.lower()
                                            def _cb(done, total, elapsed_ms):
                                                emit_progress_ex(
                                                    phase='decompress', file_type='archive',
                                                    elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                                    label=label
                                                )
                                            if is_excel_file(lower):
                                                spool_stream_to_temp_then_stream_excel(lower, dec_proc.stdout, rg_proc.stdin)
                                            elif is_csv_file(lower):
                                                stream_csv_fileobj_to_writer(dec_proc.stdout, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                            else:
                                                copy_fileobj_chunked(dec_proc.stdout, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                        except Exception:
                                            pass
                                        try:
                                            rg_proc.stdin.close()
                                        except Exception:
                                            pass
                                        try:
                                            dec_proc.stdout.close()
                                        except Exception:
                                            pass
                                        try:
                                            dec_proc.wait()
                                        except Exception:
                                            pass
                                        # 串行化：等待 rg 完成当前成员检索
                                        try:
                                            rg_proc.wait()
                                        except Exception:
                                            pass
                                    except Exception:
                                        continue
                                streamed_local = True
                            except Exception:
                                streamed_local = False
                        streamed = streamed_local
                    # 7z 优先使用 7z CLI 逐成员流式；若不可用则回退 py7zr
                    elif alc.endswith('.7z'):
                        streamed_local = False
                        if has_cmd('7z'):
                            try:
                                members = list_7z_members(full_archive)
                                for m in members:
                                    name = m.get('name')
                                    size = m.get('size')
                                    label = os.path.relpath(full_archive, data_dir) + '/' + name
                                    supports_label = check_rg_supports_label()
                                    if supports_label:
                                        cmd = rg_base.copy() + ['-a', '--label', label, '--', keyword, '-']
                                    else:
                                        cmd = rg_base.copy() + ['-a', '--', keyword, '-']
                                    try:
                                        rg_proc = subprocess.Popen(
                                            cmd,
                                            stdin=subprocess.PIPE,
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.STDOUT,
                                            shell=False,
                                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                        )
                                        extra_procs_local.append(rg_proc)
                                        all_rg_procs.append(rg_proc)
                                        if not supports_label:
                                            _proc_label_map[rg_proc.pid] = label
                                        dec_proc = subprocess.Popen(
                                            ['7z', 'x', '-so', full_archive, name],
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            shell=False,
                                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                        )
                                        extra_procs_local.append(dec_proc)
                                        try:
                                            lower = name.lower()
                                            def _cb(done, total, elapsed_ms):
                                                emit_progress_ex(
                                                    phase='decompress', file_type='archive',
                                                    elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                                    label=label
                                                )
                                            if is_excel_file(lower):
                                                spool_stream_to_temp_then_stream_excel(lower, dec_proc.stdout, rg_proc.stdin)
                                            elif is_csv_file(lower):
                                                stream_csv_fileobj_to_writer(dec_proc.stdout, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                            else:
                                                copy_fileobj_chunked(dec_proc.stdout, rg_proc.stdin, progress_cb=_cb, bytes_total=size)
                                        except Exception:
                                            pass
                                        try:
                                            rg_proc.stdin.close()
                                        except Exception:
                                            pass
                                        try:
                                            dec_proc.stdout.close()
                                        except Exception:
                                            pass
                                        try:
                                            dec_proc.wait()
                                        except Exception:
                                            pass
                                        # 串行化：等待 rg 完成本成员检索
                                        try:
                                            rg_proc.wait()
                                        except Exception:
                                            pass
                                    except Exception:
                                        continue
                                streamed_local = True
                            except Exception:
                                streamed_local = False
                        if not streamed_local:
                            try:
                                import py7zr
                                with py7zr.SevenZipFile(full_archive, mode='r') as z:
                                    names = []
                                    try:
                                        names = z.getnames()
                                    except Exception:
                                        names = []
                                    for name in names:
                                        if not name:
                                            continue
                                        label = os.path.relpath(full_archive, data_dir) + '/' + name
                                        supports_label = check_rg_supports_label()
                                        if supports_label:
                                            cmd = rg_base.copy() + ['-a', '--label', label, '--', keyword, '-']
                                        else:
                                            cmd = rg_base.copy() + ['-a', '--', keyword, '-']
                                        try:
                                            rg_proc = subprocess.Popen(
                                                cmd,
                                                stdin=subprocess.PIPE,
                                                stdout=subprocess.PIPE,
                                                stderr=subprocess.STDOUT,
                                                shell=False,
                                                preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                                            )
                                            extra_procs_local.append(rg_proc)
                                            all_rg_procs.append(rg_proc)
                                            if not supports_label:
                                                _proc_label_map[rg_proc.pid] = label
                                            # 读取单个成员数据
                                            try:
                                                dm = z.read([name])
                                                obj = dm.get(name)
                                                if obj is None:
                                                    raise Exception('py7zr read returned None')
                                                lower = name.lower()
                                                if is_excel_file(lower):
                                                    bio = io.BytesIO(obj if isinstance(obj, (bytes, bytearray)) else (obj.getvalue() if hasattr(obj, 'getvalue') else obj.read()))
                                                    spool_stream_to_temp_then_stream_excel(lower, bio, rg_proc.stdin)
                                                elif is_csv_file(lower):
                                                    if isinstance(obj, (bytes, bytearray)):
                                                        data_bytes = bytes(obj)
                                                        rg_proc.stdin.write(data_bytes)
                                                        if len(data_bytes) == 0 or data_bytes[-1] != 0x0A:
                                                            try:
                                                                rg_proc.stdin.write(b'\n')
                                                            except Exception:
                                                                pass
                                                    else:
                                                        stream_csv_fileobj_to_writer(obj, rg_proc.stdin)
                                                else:
                                                    if isinstance(obj, (bytes, bytearray)):
                                                        bio = io.BytesIO(bytes(obj))
                                                        copy_fileobj_chunked(bio, rg_proc.stdin)
                                                    else:
                                                        copy_fileobj_chunked(obj, rg_proc.stdin)
                                            except Exception:
                                                pass
                                            try:
                                                rg_proc.stdin.close()
                                            except Exception:
                                                pass
                                            total_files += 1
                                        except Exception:
                                            continue
                                streamed_local = True
                            except Exception:
                                streamed_local = False
                        streamed = streamed_local
                except Exception:
                    streamed = False

                if streamed:
                    # 已由流式处理启动 rg，无需再 extract 到临时目录
                    continue

                # 若流式失败，回退到原有解包到 temp dir 的方式（仍然存在外部 extractor 的依赖）
                try:
                    temp_dir_for_archive = tempfile.mkdtemp(prefix='rg_archive_')
                    temp_dirs.append(temp_dir_for_archive)
                    extracted_ok = False
                    try:
                        import tarfile, zipfile
                        if full_archive.lower().endswith(('.zip', '.jar', '.war')):
                            with zipfile.ZipFile(full_archive, 'r') as zf:
                                safe_extract_zip(zf, temp_dir_for_archive)
                            extracted_ok = True
                        else:
                            with tarfile.open(full_archive, 'r:*') as tf:
                                safe_extract_tar(tf, temp_dir_for_archive)
                            extracted_ok = True
                    except Exception:
                        extracted_ok = False

                    if not extracted_ok:
                        if full_archive.lower().endswith('.7z'):
                            ok, msg = try_py7zr_extract(full_archive, temp_dir_for_archive)
                            if ok:
                                extracted_ok = True
                        elif full_archive.lower().endswith('.rar'):
                            ok, msg = try_rarfile_extract(full_archive, temp_dir_for_archive)
                            if ok:
                                extracted_ok = True

                    if not extracted_ok:
                        try:
                            if full_archive.lower().endswith(('.zip', '.jar', '.war')):
                                extract_cmd = ['unzip', '-qq', full_archive, '-d', temp_dir_for_archive]
                            elif full_archive.lower().endswith(('.7z', '.rar')):
                                if has_cmd('7z'):
                                    extract_cmd = ['7z', 'x', '-y', full_archive, f'-o{temp_dir_for_archive}']
                                else:
                                    raise FileNotFoundError('7z not available')
                            else:
                                extract_cmd = ['bsdtar', '-xf', full_archive, '-C', temp_dir_for_archive]
                            extract_proc = subprocess.Popen(
                                extract_cmd,
                                stdout=subprocess.PIPE,
                                stderr=subprocess.PIPE,
                                shell=False,
                                preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None,
                                creationflags=popen_creationflags()
                            )
                            extra_procs_local.append(extract_proc)
                            extra_procs.append(extract_proc)
                            out, err = extract_proc.communicate()
                            if extract_proc.returncode != 0:
                                try:
                                    shutil.rmtree(temp_dir_for_archive)
                                except Exception:
                                    pass
                                temp_dirs.remove(temp_dir_for_archive)
                                continue
                        except FileNotFoundError:
                            try:
                                shutil.rmtree(temp_dir_for_archive)
                            except Exception:
                                pass
                            temp_dirs.remove(temp_dir_for_archive)
                            continue
                    cnt = 0
                    for _, _, fns in os.walk(temp_dir_for_archive):
                        for _ in fns:
                            cnt += 1
                    total_files += cnt if cnt > 0 else 1
                    _p_arch_dir = start_rg_for_path(temp_dir_for_archive)
                    try:
                        _p_arch_dir.wait()
                    except Exception:
                        pass
                except Exception:
                    continue

            # 为非压缩文件启动主搜索，不包括压缩文件/归档文件/Excel 文件以避免重复
            # 串行逐文件检索常规文件，确保完整枚举与稳定落盘
            for full in regular_files:
                fn = os.path.basename(full)
                fn_lower = fn.lower()
                rel_label = os.path.relpath(full, data_dir)
                def feed_fn(w, p=full, lb=rel_label, fl=fn_lower):
                    size_c = None
                    try:
                        size_c = os.path.getsize(p)
                    except Exception:
                        size_c = None
                    ft = classify_file_type(fl)
                    def _cb(done, total, elapsed_ms, ft_local=ft, lb_local=lb):
                        emit_progress_ex(
                            phase='scan', file_type=ft_local,
                            elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                            label=lb_local
                        )
                    try:
                        with open(p, 'rb') as f:
                            copy_fileobj_chunked(f, w, progress_cb=_cb, bytes_total=size_c)
                    except Exception:
                        pass
                _p_reg = start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                try:
                    _p_reg.wait()
                except Exception:
                    pass

            # 为每个压缩文件（流式）启动 rg。优先使用外部工具；如果缺失，则使用 Python 流式输入
            for full in compressed_files:
                fn = os.path.basename(full)
                fn_lower = fn.lower()
                dec_cmd = build_decompress_command(fn_lower, full)
                rel_label = os.path.relpath(full, data_dir)
                if dec_cmd:
                    try:
                        decompressor_proc = subprocess.Popen(
                            dec_cmd,
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE,
                            shell=False,
                            preexec_fn=os.setpgrp if hasattr(os, "setpgrp") else None
                        )
                        extra_procs_local.append(decompressor_proc)
                        extra_procs.append(decompressor_proc)
                        inner_lower = strip_single_compress_ext(fn_lower)
                        if is_excel_file(inner_lower):
                            def feed_fn(w, proc=decompressor_proc, name=inner_lower, lb=rel_label, orig=full):
                                def _cb(done, total, elapsed_ms):
                                    emit_progress_ex(
                                        phase='decompress', file_type='compressed',
                                        elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                        label=lb
                                    )
                                # 目录扫描下的单压缩 Excel：直接流到临时文件再转换
                                spool_stream_to_temp_then_stream_excel(name, proc.stdout, w)
                            _p_cf = start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                            try:
                                _p_cf.wait()
                            except Exception:
                                pass
                        else:
                            def feed_fn(w, proc=decompressor_proc, lb=rel_label, orig=full):
                                size_c = None
                                try:
                                    size_c = os.path.getsize(orig)
                                except Exception:
                                    size_c = None
                                def _cb(done, total, elapsed_ms):
                                    emit_progress_ex(
                                        phase='decompress', file_type='compressed',
                                        elapsed_ms=elapsed_ms, bytes_done=done, bytes_total=total,
                                        label=lb
                                    )
                                copy_fileobj_chunked(proc.stdout, w, progress_cb=_cb, bytes_total=size_c)
                            _p_cf2 = start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                            try:
                                _p_cf2.wait()
                            except Exception:
                                pass
                    except FileNotFoundError:
                        inner_lower = strip_single_compress_ext(fn_lower)
                        if is_excel_file(inner_lower):
                            def feed_fn(w, p=full, e=fn_lower, inner=inner_lower):
                                bio = io.BytesIO()
                                python_decompress_feed(p, e, bio)
                                data = bio.getvalue()
                                stream_excel_bytes_to_writer(inner, data, w)
                            _p_py1 = start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                            try:
                                _p_py1.wait()
                            except Exception:
                                pass
                        else:
                            def feed_fn(w, p=full, e=fn_lower):
                                python_decompress_feed(p, e, w)
                            _p_py2 = start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                            try:
                                _p_py2.wait()
                            except Exception:
                                pass
                else:
                    inner_lower = strip_single_compress_ext(fn_lower)
                    if is_excel_file(inner_lower):
                        def feed_fn(w, p=full, e=fn_lower, inner=inner_lower):
                            bio = io.BytesIO()
                            python_decompress_feed(p, e, bio)
                            data = bio.getvalue()
                            stream_excel_bytes_to_writer(inner, data, w)
                        try:
                            start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                        except Exception:
                            continue
                    else:
                        def feed_fn(w, p=full, e=fn_lower):
                            python_decompress_feed(p, e, w)
                        try:
                            start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                        except Exception:
                            continue

            # 新增：对每个 excel 文件做流式转换并交给 rg（不落盘）
            for full in excel_files:
                fn = os.path.basename(full)
                rel_label = os.path.relpath(full, data_dir)
                # python feed：把 excel 转为文本写入 rg stdin
                def feed_fn(w, p=full):
                    stream_excel_to_writer(p, w)
                try:
                    _p_excel2 = start_rg_for_path('-', label=rel_label, python_stream_feed=feed_fn)
                    try:
                        _p_excel2.wait()
                    except Exception:
                        pass
                except Exception:
                    continue

        # 如果没有启动 RG，出错
        if not all_rg_procs:
            socketio.emit('message', {'message': 'No searchable files or rg failed to start\n'})
            return "No search procs"

        # 统一监听这些 rg 进程的 stdout：队列已在前文创建，并在每次 rg 启动时开启前向线程
        total_procs = len(all_rg_procs)

        # 将第一个 rg 进程作为主引用（用于 cancel 等），但后台会管理所有进程
        proc = all_rg_procs[0]

    except Exception as e:
        socketio.emit('message', {'message': f'Start failed: {e}\n'})
        for p in extra_procs_local:
            try:
                os.killpg(os.getpgid(p.pid), signal.SIGTERM)
            except Exception:
                try:
                    p.terminate()
                except Exception:
                    pass
        for d in temp_dirs:
            try:
                shutil.rmtree(d)
            except Exception:
                pass
        extra_procs_local = []
        proc = None
        return "Error"

    # 读取与解析合并后的流（从队列读取）
    request_start_ns = time.perf_counter_ns()
    def get_output_loop():
        global proc, extra_procs, temp_dirs, _proc_label_map
        # 使用关键字生成导出流标识；导出文件已在提交阶段创建
        safe_kw = _sanitize_keyword(keyword)
        # match_count 现在按实际输出的“内容区块”计数，保证与页面显示一致
        match_count = 0
        files_done = 0
        try:
            import json
            with app.app_context():
                context_before_n = context_before if context_before > 0 else 0
                context_after_n = context_after if context_after > 0 else 0
                before_lines = []
                after_lines = []
                block_main = None
                block_ready = False
                after_emit_count = 0
                first_block = True
                # 搜索阶段计时（毫秒级）
                search_start_ns = {}
                last_progress_tick_ns = 0
                # 记录每个进程当前文件的可读标签（路径或显式 label），便于进度与写盘头信息
                owner_current_label = {}

                # 发送初始匹配数和总文件数（files_total）
                try:
                    elapsed_ms_total = int((time.perf_counter_ns() - request_start_ns) / 1_000_000)
                except Exception:
                    elapsed_ms_total = 0
                emit_progress_ex(matches=match_count, files_total=total_files, files_done=files_done, elapsed_ms=elapsed_ms_total)

                # loop: 从队列中读取行，直到所有 rg 进程发出 EOF 标记
                eof_set = set()
                while True:
                    if cancel_requested:
                        break
                    try:
                        raw_item, owner = q.get(timeout=0.05)  # Reduced timeout for more frequent cancel checks
                    except queue.Empty:
                        # 更频繁地检查取消情况
                        if cancel_requested:
                            break
                        # 检查进程是否全部退出并且队列空 => 结束
                        all_ended = True
                        for p in extra_procs_local:
                            if p.poll() is None:
                                all_ended = False
                                break
                        if all_ended and q.empty():
                            break
                        # 空闲心跳：周期性推送耗时与累计状态（约每200ms一次）
                        try:
                            now_ns = time.perf_counter_ns()
                            if last_progress_tick_ns == 0 or (now_ns - last_progress_tick_ns) >= 200_000_000:
                                elapsed_ms_total = int((now_ns - request_start_ns) / 1_000_000)
                                emit_progress_ex(matches=match_count, files_total=total_files, files_done=files_done, elapsed_ms=elapsed_ms_total)
                                last_progress_tick_ns = now_ns
                        except Exception:
                            pass
                        continue

                    if raw_item is None:
                        eof_set.add(owner)
                        if len(eof_set) >= total_procs:
                            break
                        else:
                            continue

                    # 解析单行
                    try:
                        line = raw_item.decode('utf-8', errors='replace')
                    except Exception:
                        continue
                    if not line.strip():
                        continue
                    try:
                        obj = json.loads(line)
                    except Exception:
                        # 非 JSON 行（通常为外部解压/提示），忽略以保持纯净内容输出
                        continue

                    typ = obj.get('type')
                    # 处理 'begin' 以更新 files_done
                    if typ == 'begin':
                        # 在处理之前检查是否已取消
                        if cancel_requested:
                            break
                        # 如果 rg 不支持 --label，则尝试注入 label 到 obj 的 path 字段（仅本地使用）
                        if 'data' in obj and 'path' not in obj.get('data', {}) and owner in _proc_label_map:
                            try:
                                obj['data']['path'] = {'text': _proc_label_map.get(owner)}
                            except Exception:
                                pass
                        # 搜索开始计时
                        try:
                            search_start_ns[owner] = time.perf_counter_ns()
                        except Exception:
                            pass
                        # 写入文件头到导出（无论是否会命中），保证实时落盘记录
                        try:
                            data_path = (obj.get('data') or {}).get('path') or {}
                            label_text = data_path.get('text') if isinstance(data_path, dict) else None
                        except Exception:
                            label_text = None
                        if not label_text:
                            label_text = _proc_label_map.get(owner)
                        if label_text:
                            owner_current_label[owner] = label_text
                            append_export_text(safe_kw, f"[{label_text}]\n")
                        # 本次文件开始时不增加 files_done，保持以文件结束为准
                        # 重置本文件缓冲
                        before_lines = []
                        after_lines = []
                        block_main = None
                        block_ready = False
                    elif typ == 'context':
                        # 在处理之前检查是否已取消
                        if cancel_requested:
                            break
                        data = obj.get('data', {})
                        line_text = (data.get('lines', {}).get('text', '') or '').strip()
                        if not block_ready:
                            before_lines.append(line_text)
                            if len(before_lines) > context_before_n:
                                before_lines.pop(0)
                        else:
                            # 实时输出后文行
                            append_export_text(safe_kw, line_text + '\n')
                            emit_message_utf(line_text + '\n')
                            after_emit_count += 1
                            # 若后文满足要求，完成该块
                            if after_emit_count >= context_after_n:
                                match_count += 1
                                try:
                                    elapsed_ms_total = int((time.perf_counter_ns() - request_start_ns) / 1_000_000)
                                except Exception:
                                    elapsed_ms_total = 0
                                emit_progress_ex(matches=match_count, files_total=total_files, files_done=files_done, elapsed_ms=elapsed_ms_total)
                                first_block = False
                                block_ready = False
                                # 将主行加入前文缓冲，供后续块使用
                                before_lines.append(block_main)
                                if len(before_lines) > context_before_n:
                                    before_lines.pop(0)
                                block_main = None
                                after_emit_count = 0
                                after_lines = []
                    elif typ == 'match':
                        # 在处理之前检查是否已取消
                        if cancel_requested:
                            break
                        data = obj.get('data', {})
                        line_text = (data.get('lines', {}).get('text', '') or '').strip()

                        # 若上一块仍未结束，立即补齐后文并完成它
                        if block_ready:
                            while after_emit_count < context_after_n:
                                append_export_text(safe_kw, '\n')
                                emit_message_utf('\n')
                                after_emit_count += 1
                            match_count += 1
                            try:
                                elapsed_ms_total = int((time.perf_counter_ns() - request_start_ns) / 1_000_000)
                            except Exception:
                                elapsed_ms_total = 0
                            emit_progress_ex(matches=match_count, files_total=total_files, files_done=files_done, elapsed_ms=elapsed_ms_total)
                            first_block = False
                            block_ready = False

                        # 开始新块：分隔空行、输出前文与主行并实时写盘
                        if not first_block:
                            append_export_text(safe_kw, '\n')
                            emit_message_utf('\n')
                        for i in range(context_before_n):
                            t = before_lines[i] if i < len(before_lines) else ''
                            append_export_text(safe_kw, t + '\n')
                            emit_message_utf(t + '\n')
                        append_export_text(safe_kw, line_text + '\n')
                        emit_message_utf(line_text + '\n')

                        # 标记进入后文阶段
                        block_main = line_text
                        after_emit_count = 0
                        after_lines = []
                        block_ready = True
                    elif typ == 'end':
                        # 文件边界：若当前块仍未完成，补齐后文空行并完成该块
                        if block_ready:
                            while after_emit_count < context_after_n:
                                append_export_text(safe_kw, '\n')
                                emit_message_utf('\n')
                                after_emit_count += 1
                            match_count += 1
                            try:
                                elapsed_ms_total = int((time.perf_counter_ns() - request_start_ns) / 1_000_000)
                            except Exception:
                                elapsed_ms_total = 0
                            emit_progress_ex(matches=match_count, files_total=total_files, files_done=files_done, elapsed_ms=elapsed_ms_total)
                            first_block = False
                            block_ready = False
                            block_main = None
                            after_emit_count = 0
                            after_lines = []
                        # 在文件末尾重置缓冲区
                        before_lines = []
                        after_lines = []
                        block_main = None
                        block_ready = False
                        # 文件结束：更新完成数并推送进度
                        files_done += 1
                        socketio.emit('progress', {
                            'matches': match_count,
                            'files_total': total_files,
                            'files_done': files_done
                        })
                        # 搜索结束计时并发出扩展进度（毫秒级）
                        try:
                            ns = search_start_ns.pop(owner, None)
                            if ns is not None:
                                elapsed_ms = int((time.perf_counter_ns() - ns) / 1_000_000)
                                # 推测文件类型：优先使用当前标签（路径），再基于后缀分类
                                label = owner_current_label.get(owner) or _proc_label_map.get(owner)
                                ft = classify_file_type(label.lower()) if label else None
                                emit_progress_ex(
                                    phase='search_end', file_type=ft,
                                    elapsed_ms=elapsed_ms, files_total=total_files,
                                    files_done=files_done, label=label
                                )
                        except Exception:
                            pass

                    # 逐行流式写盘：已在 match/context/end 分支实时输出并更新进度
        except Exception:
            # 兜底：管道可能已断，尝试将剩余 stdout 简单转发
            with app.app_context():
                try:
                    for p in extra_procs_local:
                        if p and p.stdout:
                            for raw in p.stdout:
                                if cancel_requested:
                                    break
                                text = raw.decode('utf-8', errors='replace')
                                # 补上 label（如果有）
                                label = _proc_label_map.get(getattr(p, 'pid', None))
                                if label:
                                    text = f"[{label}] {text}"
                                append_export_text(safe_kw, text)
                                emit_message_utf(text)
                except Exception:
                    pass
        finally:
            # 清理：等待所有 rg 及解压进程退出，终止额外进程并清理临时目录
            try:
                for p in extra_procs_local:
                    try:
                        # 给进程更多时间自然完成，特别是对于7z等大文件处理
                        p.wait(timeout=5.0)
                    except subprocess.TimeoutExpired:
                        # 如果进程超时，再尝试温和终止
                        try:
                            if hasattr(p, 'terminate'):
                                p.terminate()
                                p.wait(timeout=2.0)
                        except Exception:
                            pass
                    except Exception:
                        pass
            except Exception:
                pass

            # 只有真正无法结束的进程才强制终止
            for p in list(extra_procs_local):
                try:
                    if hasattr(p, 'poll') and p.poll() is None:
                        # 进程仍在运行，强制终止
                        try:
                            os.killpg(os.getpgid(p.pid), signal.SIGTERM)
                        except Exception:
                            try:
                                p.terminate()
                            except Exception:
                                pass
                except Exception:
                    pass

            # 合并到全局 extra_procs 便于 cancel
            for p in extra_procs_local:
                if p not in extra_procs:
                    extra_procs.append(p)

            # 删除临时目录（若有）
            for d in list(temp_dirs):
                try:
                    shutil.rmtree(d)
                except Exception:
                    pass
            temp_dirs = []

            # 清理 pid->label 映射（避免内存泄漏）
            _proc_label_map = {}

            # 置空全局 proc（注意：proc 可能是第一个 rg）
            proc = None

            # 发送最终匹配数一次，确保前端能收到最终值
            try:
                elapsed_ms_total = int((time.perf_counter_ns() - request_start_ns) / 1_000_000)
            except Exception:
                elapsed_ms_total = 0
            emit_progress_ex(matches=match_count, files_total=total_files, files_done=files_done, elapsed_ms=elapsed_ms_total)
            # 已采用流式写盘，关闭导出流
            close_export_stream(safe_kw)

            # 确保所有进程都已完成，特别是对于7z等大文件处理
            try:
                time.sleep(0.5)
            except Exception:
                pass

            # 在发送完成消息前，再次发送最终匹配数，确保前端收到
            try:
                elapsed_ms_total = int((time.perf_counter_ns() - request_start_ns) / 1_000_000)
            except Exception:
                elapsed_ms_total = 0
            emit_progress_ex(matches=match_count, files_total=total_files, files_done=files_done, elapsed_ms=elapsed_ms_total)

            with app.app_context():
                if not cancel_requested:
                    emit_message_utf('Done\n')

    # 启动后台读写线程
    thread = threading.Thread(target=get_output_loop, daemon=True)
    thread.start()
    # 立即通知前端搜索启动（前端会显示进度条）并发送初始 total_files
    emit_message_utf('Started\n')
    try:
        elapsed_ms_total = int((time.perf_counter_ns() - request_start_ns) / 1_000_000)
    except Exception:
        elapsed_ms_total = 0
    emit_progress_ex(files_total=total_files, files_done=0, matches=0, elapsed_ms=elapsed_ms_total)
    return "Started"


@app.route('/cancel', methods=['POST'])
def cancel():
    global proc, extra_procs, temp_dirs, _proc_label_map, cancel_requested
    request_cancel_ns = time.perf_counter_ns()
    cancel_requested = True
    
    # 立即发送取消消息以通知前端
    emit_message_utf('Cancelled\n')
    emit_progress_ex(phase='cancelled')

    # 安全关闭流的辅助函数
    def _close_streams(p):
        try:
            if p:
                try:
                    s = getattr(p, 'stdin', None)
                    if s: s.close()
                except Exception:
                    pass
                try:
                    s = getattr(p, 'stdout', None)
                    if s: s.close()
                except Exception:
                    pass
                try:
                    s = getattr(p, 'stderr', None)
                    if s: s.close()
                except Exception:
                    pass
        except Exception:
            pass

    # 强制终止进程（跨平台，优先关闭管道，其次温和终止，最后强杀）
    def _terminate_proc(p):
        if not p:
            return
        # 先关闭流以打破阻塞，促使子进程尽快退出
        _close_streams(p)
        try:
            if hasattr(p, 'poll') and p.poll() is None:
                if os.name != 'nt':
                    try:
                        # 请先尝试终止进程组
                        pgid = os.getpgid(p.pid)
                        os.killpg(pgid, signal.SIGTERM)
                    except Exception:
                        try:
                            p.terminate()
                        except Exception:
                            pass
                else:
                    # Windows：先尝试 terminate，若仍存活则使用 taskkill 强制终止进程树
                    try:
                        p.terminate()
                    except Exception:
                        pass
                
                # 稍等片刻以优雅终止
                try:
                    p.wait(timeout=0.5)
                except Exception:
                    pass
                
                # 如果仍在运行，强制结束
                if hasattr(p, 'poll') and p.poll() is None:
                    if os.name == 'nt':
                        try:
                            subprocess.run(['taskkill', '/PID', str(p.pid), '/T', '/F'],
                                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, shell=False)
                        except Exception:
                            pass
                    else:
                        try:
                            # 先尝试终止进程组，然后再终止单个进程
                            try:
                                pgid = os.getpgid(p.pid)
                                os.killpg(pgid, signal.SIGKILL)
                            except Exception:
                                p.kill()
                        except Exception:
                            pass
        except Exception:
            pass
        # 再次尝试关闭流（终止后）
        _close_streams(p)

    # 1) 首先终止额外子进程（解压器等），然后终止主检索进程
    # 这样可以更快地停止数据流
    for p in list(extra_procs):
        _terminate_proc(p)
    extra_procs = []
    
    # 2) 终止主检索进程
    _terminate_proc(proc)

    # 3) 清理临时目录
    for d in list(temp_dirs):
        try:
            shutil.rmtree(d)
        except Exception:
            pass
    temp_dirs = []

    # 4) 关闭并清空导出流，避免资源泄漏（已流式写盘）
    close_all_export_streams()

    # 5) 状态复位
    proc = None
    _proc_label_map = {}

    # 6) 通知前端取消
    emit_message_utf('Cancelled\n')
    try:
        elapsed_ms_total = int((time.perf_counter_ns() - request_cancel_ns) / 1_000_000)
    except Exception:
        elapsed_ms_total = 0
    emit_progress_ex(phase='cancelled', elapsed_ms=elapsed_ms_total)
    return jsonify({"status": "cancelled"}), 200


@app.route('/download')
def download():
    keyword = request.args.get('keyword')
    if not keyword:
        abort(400)
    safe = ''.join(c for c in keyword if c.isalnum() or c in (' ', '_', '-')).strip()
    if not safe:
        safe = 'search'
    # 与写入路径保持一致：自适应导出目录
    exports_dir = get_exports_dir()
    candidates = []
    if os.path.isdir(exports_dir):
        for fn in os.listdir(exports_dir):
            if fn.startswith(safe):
                candidates.append(fn)
    if not candidates:
        abort(404)
    candidates.sort(key=lambda n: os.path.getmtime(os.path.join(exports_dir, n)), reverse=True)
    return send_from_directory(exports_dir, candidates[0], as_attachment=True)


@app.route('/files')
def list_files():
    data_dir = '/data'
    if not os.path.isdir(data_dir):
        data_dir = os.path.dirname(__file__)
    files = []
    try:
        for root, _, fns in os.walk(data_dir):
            for fn in fns:
                # 仅返回相对路径以便 UI 选择（包含压缩/归档文件）
                full = os.path.join(root, fn)
                rel = os.path.relpath(full, data_dir)
                files.append(rel)
    except Exception:
        files = []
    return jsonify(files)

if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=5000)
